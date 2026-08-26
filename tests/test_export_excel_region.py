"""
ปุ่มดาวน์โหลด Excel ต้องให้ตัวเลขชุดเดียวกับ Excel ที่ขั้นกระจายสร้าง

ขั้นกระจายแก้ไปแล้วว่าโหมดรวมภาคต้องใช้ "ไฟล์เป้ารวม" เป็นแหล่งของแถวเป้าหีบ
(optimize._excel_target_boxes_path) แต่ปุ่มดาวน์โหลดยังชี้ไฟล์ของทีมเดียวเสมอ
Excel สองไฟล์ของรอบเดียวกันจึงไม่ตรงกัน: ไฟล์จากปุ่มนี้เอาเป้าของทีมเดียวมาคู่กับ
หีบของทั้งภาค ดูเหมือนกระจายเกินเป้ามหาศาล · และการเติมราคา/ชื่อสินค้าก็อ่านจาก
ไฟล์ทีมเดียว SKU ที่ทีมนั้นไม่มีจึงได้ราคา 0 มูลค่ารวมในไฟล์เลยต่ำกว่าจริงเป็นก้อน
"""

import os
import shutil
import tempfile
import unittest

import openpyxl
import pandas as pd

from backend.core.paths import (
    target_boxes_cache_path,
    target_boxes_union_cache_path,
)
from backend.generate_excel import create_target_excel
from backend.services.exporting import _export_target_boxes_path

MONTH, YEAR = 9, 2026
SUP = "SL397"


class TestExportTargetBoxesSource(unittest.TestCase):
    def setUp(self):
        self._cwd = os.getcwd()
        self._tmpdir = tempfile.mkdtemp(prefix="export_region_")
        os.makedirs(os.path.join(self._tmpdir, "data"), exist_ok=True)
        os.chdir(self._tmpdir)

    def tearDown(self):
        os.chdir(self._cwd)
        shutil.rmtree(self._tmpdir, ignore_errors=True)

    def _write(self, path: str, boxes: int) -> str:
        pd.DataFrame([{
            "sku": "734046", "price_per_box": 352.0,
            "supervisor_target_boxes": boxes,
        }]).to_csv(path, index=False)
        return path

    def test_falls_back_to_own_team_file_when_there_is_no_union(self):
        own = self._write(target_boxes_cache_path(SUP, MONTH, YEAR), 100)
        self.assertEqual(_export_target_boxes_path(SUP, MONTH, YEAR), own)

    def test_uses_the_union_file_after_a_region_wide_run(self):
        self._write(target_boxes_cache_path(SUP, MONTH, YEAR), 100)
        union = self._write(target_boxes_union_cache_path(SUP, MONTH, YEAR), 450)
        os.utime(union, (os.path.getmtime(union) + 60,) * 2)
        self.assertEqual(_export_target_boxes_path(SUP, MONTH, YEAR), union)

    def test_a_later_single_team_run_wins_again(self):
        """กระจายรวมภาคแล้วกลับมากระจายทีมเดียว ไฟล์รวมภาคเก่าต้องไม่ถูกหยิบมาใช้อีก"""
        union = self._write(target_boxes_union_cache_path(SUP, MONTH, YEAR), 450)
        own = self._write(target_boxes_cache_path(SUP, MONTH, YEAR), 100)
        os.utime(own, (os.path.getmtime(union) + 60,) * 2)
        self.assertEqual(_export_target_boxes_path(SUP, MONTH, YEAR), own)

    def test_without_a_period_it_is_the_old_behaviour(self):
        path = _export_target_boxes_path(SUP, None, None)
        self.assertTrue(path.endswith("target_boxes.csv"), path)


class TestRegionExcelFile(unittest.TestCase):
    """สร้างไฟล์จริงแล้วเปิดอ่าน — ผลกระจายรวมภาคต้องครบทุกทีมและอ่านออกว่าเป็นก้อนไหน"""

    def setUp(self):
        self._cwd = os.getcwd()
        self._tmpdir = tempfile.mkdtemp(prefix="excel_file_")
        os.makedirs(os.path.join(self._tmpdir, "data"), exist_ok=True)
        os.chdir(self._tmpdir)

    def tearDown(self):
        os.chdir(self._cwd)
        shutil.rmtree(self._tmpdir, ignore_errors=True)

    def _build(self, scope_sup_ids):
        rows = []
        # พนักงานคนละทีมกัน อยู่ในผลกระจายก้อนเดียวกัน (โหมดรวมภาค)
        for emp, boxes in (("E1", 60), ("E2", 40), ("E3", 50)):
            rows.append({
                "emp_id": emp, "sku": "734046", "allocated_boxes": boxes,
                "hist_avg": 10.0, "hist_ly_same_month": 0.0, "hist_prev_month": 0.0,
                "price_per_box": 352.0, "brand_name_thai": "ปรุงทิพย์",
                "brand_name_english": "", "product_name_thai": "", "product_name_english": "",
            })
        result_csv = "data/result_test.csv"
        pd.DataFrame(rows).to_csv(result_csv, index=False)
        target_csv = "data/target_boxes_union.csv"
        pd.DataFrame([{
            "sku": "734046", "price_per_box": 352.0, "supervisor_target_boxes": 150,
            "brand_name_thai": "ปรุงทิพย์",
        }]).to_csv(target_csv, index=False)
        out = create_target_excel(
            result_csv=result_csv,
            output_path="data/out.xlsx",
            brand_filter="ALL",
            yellow_map={"E1": 21120.0, "E2": 14080.0, "E3": 17600.0},
            sup_id="SL397",
            target_boxes_csv=target_csv,
            scope_sup_ids=scope_sup_ids,
        )
        self.assertIsNotNone(out, "ต้องสร้างไฟล์ได้")
        wb = openpyxl.load_workbook(out)
        return wb.active

    def test_region_run_labels_the_scope_on_the_header(self):
        ws = self._build(["SL397", "SL460", "SL509"])
        title = str(ws.cell(row=1, column=1).value or "")
        self.assertIn("รวม 3 ทีม", title)
        self.assertIn("SL460", title)

    def test_single_team_header_is_unchanged(self):
        ws = self._build([])
        title = str(ws.cell(row=1, column=1).value or "")
        self.assertIn("Supervisor: SL397", title)
        self.assertNotIn("รวม", title.split("แบรนด์")[0].replace("Supervisor", ""))

    def test_every_employee_in_the_region_is_in_the_file(self):
        """หัวใจของ 'ดาวน์โหลดรวมทั้งภาค' — ต้องไม่หล่นพนักงานของทีมอื่น"""
        ws = self._build(["SL397", "SL460", "SL509"])
        seen = {
            str(c.value).strip()
            for row in ws.iter_rows()
            for c in row
            if c.value is not None
        }
        for emp in ("E1", "E2", "E3"):
            self.assertTrue(
                any(emp in v for v in seen), f"ไม่พบพนักงาน {emp} ในไฟล์"
            )


if __name__ == "__main__":
    unittest.main()
