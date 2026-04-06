"""
main.py — Target Allocation API (v3 — Production)
────────────────────────────────────────────────────────────────────
uvicorn main:app --reload

การไหลของข้อมูล:
  1. GET  /data/employees  → ดึงพนักงาน + SKU + LY + 3M hist จาก Fabric
  2. POST /optimize        → OR engine ตาม strategy
  3. POST /export/excel    → สร้าง Excel (กรองรายแบรนด์ได้)
  4. GET  /download/excel  → ดาวน์โหลดไฟล์ที่สร้าง
"""

import os
import re
import random
import logging
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
import json

from fastapi import FastAPI, HTTPException, Query, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
import pandas as pd
try:
    import httpx as _httpx
    _HTTPX_OK = True
except ImportError:
    _HTTPX_OK = False
import io

from OR_engine import allocate_boxes
from generate_excel import create_target_excel
from fabric_dax_connector import FabricDAXConnector

# ── Logging ──────────────────────────────────────────────
os.makedirs("data", exist_ok=True) 
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("data/app.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger("target_allocation")

# ── Startup helpers (ต้อง define ก่อน app เพราะ FastAPI ต้องการ lifespan ตอน init) ──
def _cleanup_old_caches(max_age_days: int = 7):
    cutoff = datetime.now() - timedelta(days=max_age_days)
    try:
        for fname in os.listdir("data"):
            if fname.startswith(("hist_cache_", "emp_cache_")):
                fpath = os.path.join("data", fname)
                if datetime.fromtimestamp(os.path.getmtime(fpath)) < cutoff:
                    os.remove(fpath)
                    logger.info("Cleaned old cache: %s", fname)
    except Exception as e:
        logger.warning("Cache cleanup error: %s", e)

@asynccontextmanager
async def lifespan(app_: FastAPI):
    os.makedirs("data", exist_ok=True)
    _cleanup_old_caches(max_age_days=7)
    yield

# ── App ───────────────────────────────────────────────────
app = FastAPI(title="Target Allocation API", version="3.0", lifespan=lifespan)

# CORS — allow_credentials=False เพราะไม่ได้ใช้ cookie/session
# ⚠️ allow_credentials=True + allow_origins=["*"] ผิด HTTP spec
#    Starlette จะ drop CORS headers เงียบๆ → browser block ทุก request จาก file://
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Known supervisors ─────────────────────────────────────
# โหลดจาก env var ก่อน — fallback ไปค่า hardcoded เพื่อกัน startup crash
# ตั้งค่า: set KNOWN_MANAGERS=SL330,SL374,SL999,SL001 ใน start_server.bat
_km_env = os.environ.get("KNOWN_MANAGERS", "")
KNOWN_MANAGERS = [m.strip() for m in _km_env.split(",") if m.strip()] or ["SL330", "SL374", "SL999"]

# ── Fallback price per SKU ────────────────────────────────
PRICE_FALLBACK = {
    "624007": 240.00,
    "624015": 212.00,
    "624049": 335.00,
    "624056": 290.00,
    "624114": 212.00,
    "624163": 232.71,
}

VALID_STRATEGIES = ("L3M", "L6M", "EVEN", "PUSH", "LP")

# ── API key สำหรับ /translate/strategy ───────────────────
# ใช้ OpenRouter (แนะนำ) หรือ Anthropic โดยตรง
# OpenRouter:  set OPENROUTER_API_KEY=sk-or-...  ใน start_server.bat
# Anthropic:   set ANTHROPIC_API_KEY=sk-ant-...  ใน start_server.bat
_OPENROUTER_API_KEY = os.environ.get("OPENROUTER_API_KEY", "")
_ANTHROPIC_API_KEY  = os.environ.get("ANTHROPIC_API_KEY", "")


# ══════════════════════════════════════════════════════════
#  POST /translate/strategy — proxy AI API สำหรับ Custom Strategy
#  รองรับทั้ง OpenRouter (แนะนำ) และ Anthropic โดยตรง
# ══════════════════════════════════════════════════════════
class TranslateStrategyRequest(BaseModel):
    text: str = Field(..., min_length=1, max_length=500)
    emp_count: int = Field(default=10, ge=1)
    sku_count: int = Field(default=20, ge=1)

@app.post("/translate/strategy")
async def translate_strategy(req: TranslateStrategyRequest):
    import json as _json
    if not _HTTPX_OK:
        raise HTTPException(503, detail="httpx ไม่ได้ติดตั้ง — รัน: pip install httpx แล้ว restart server")

    system_prompt = f"""คุณเป็น OR expert ที่แปลง natural language เป็น allocation parameter
ตอบเป็น JSON เท่านั้น ไม่มี markdown ไม่มีคำอธิบายนอก JSON

Parameter ที่ใช้ได้:
- base_strategy: "L3M" | "L6M" | "EVEN" | "PUSH" (เลือก 1 ที่ใกล้เคียงที่สุด)
- cap_multiplier: number 1.5-5.0 (จำกัดสูงสุดเท่าค่าเฉลี่ย × cap)
  * 1.5 = เกลี่ยใกล้เคียงมาก  2.0 = ยืดหยุ่นพอดี  3.0 = ค่อนข้างอิสระ  5.0 = เกือบไม่ cap
- description_th: อธิบายสั้นๆ (ภาษาไทย ≤ 40 ตัวอักษร)
- confidence: "high" | "medium" | "low"
- note: คำเตือนถ้ามี (optional)

บริบท: ทีมมี {req.emp_count} คน, {req.sku_count} SKU"""

    try:
        async with _httpx.AsyncClient(timeout=20) as client:

            # ── OpenRouter (แนะนำ — ราคาถูก ใช้ได้หลาย model) ──
            if _OPENROUTER_API_KEY:
                resp = await client.post(
                    "https://openrouter.ai/api/v1/chat/completions",
                    headers={
                        "Authorization": f"Bearer {_OPENROUTER_API_KEY}",
                        "Content-Type": "application/json",
                    },
                    json={
                        "model": "anthropic/claude-haiku-4-5",  # เปลี่ยน model ได้เลย
                        "max_tokens": 300,
                        "messages": [
                            {"role": "system", "content": system_prompt},
                            {"role": "user",   "content": req.text},
                        ],
                    },
                )
                if resp.status_code != 200:
                    raise HTTPException(502, detail=f"OpenRouter error: {resp.status_code} — {resp.text[:200]}")
                raw = resp.json()["choices"][0]["message"]["content"]

            # ── Anthropic โดยตรง (fallback) ──
            elif _ANTHROPIC_API_KEY:
                resp = await client.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={
                        "x-api-key": _ANTHROPIC_API_KEY,
                        "anthropic-version": "2023-06-01",
                        "content-type": "application/json",
                    },
                    json={
                        "model": "claude-haiku-4-5-20251001",
                        "max_tokens": 300,
                        "system": system_prompt,
                        "messages": [{"role": "user", "content": req.text}],
                    },
                )
                if resp.status_code != 200:
                    raise HTTPException(502, detail=f"Anthropic API error: {resp.status_code}")
                raw = next(
                    (c["text"] for c in resp.json().get("content", []) if c.get("type") == "text"), ""
                )

            else:
                raise HTTPException(
                    503,
                    detail="ยังไม่ได้ตั้งค่า API key — ใส่ OPENROUTER_API_KEY หรือ ANTHROPIC_API_KEY ใน start_server.bat แล้ว restart"
                )

        clean = raw.replace("```json", "").replace("```", "").strip()
        try:
            params = _json.loads(clean)
        except _json.JSONDecodeError as je:
            logger.warning("translate_strategy: JSON parse error: %s | raw=%s", je, clean[:200])
            raise HTTPException(500, detail="AI ตอบกลับไม่ถูกรูปแบบ JSON — ลองพิมพ์ใหม่อีกครั้ง")

        valid_strats = ["L3M", "L6M", "EVEN", "PUSH"]
        if params.get("base_strategy") not in valid_strats:
            params["base_strategy"] = "L3M"
        params["cap_multiplier"] = min(5.0, max(1.5, float(params.get("cap_multiplier", 3.0))))
        # ตรวจ required fields — ป้องกัน frontend crash
        params.setdefault("description_th", "Custom strategy")
        params.setdefault("confidence", "medium")
        params.setdefault("note", None)
        return params

    except _httpx.TimeoutException:
        raise HTTPException(504, detail="AI API timeout — ลองใหม่อีกครั้ง")
    except HTTPException:
        raise
    except Exception as e:
        logger.error("translate_strategy error: %s", e)
        raise HTTPException(500, detail=str(e))


# ══════════════════════════════════════════════════════════
#  Pydantic models
# ══════════════════════════════════════════════════════════
class YellowTargetInput(BaseModel):
    emp_id: str
    yellow_target: float = Field(ge=0)

class LockedEditInput(BaseModel):  # ข้อ 11
    emp_id: str
    sku: str
    locked_boxes: int = Field(ge=0)

class OptimizeRequest(BaseModel):
    yellowTargets: list[YellowTargetInput]
    strategy: str = "L3M"
    force_min_one: bool = False
    locked_edits: list[LockedEditInput] = []
    cap_multiplier: float | None = None  # Custom strategy override (1.5–5.0)

class AllocationRow(BaseModel):
    emp_id: str
    sku: str
    allocated_boxes: int = Field(ge=0)
    hist_avg: float = 0.0
    price_per_box: float = 0.0
    brand_name_thai: str = ""
    brand_name_english: str = ""
    product_name_thai: str = ""

class ExportRequest(BaseModel):
    allocations: list[AllocationRow]
    brand_filter: str = "ALL"
    yellow_targets: list[YellowTargetInput] = []


# ══════════════════════════════════════════════════════════
#  Helpers: safe path builders (prevent path traversal)
# ══════════════════════════════════════════════════════════
def _safe_id(s: str) -> str:
    """Sanitize sup_id / strategy สำหรับใส่ใน filename"""
    return re.sub(r"[^A-Za-z0-9_]", "_", str(s))

def hist_cache_path(sup_id: str, month: int, year: int) -> str:
    return f"data/hist_cache_{_safe_id(sup_id)}_{year}_{month:02d}.csv"

def emp_cache_path(sup_id: str, month: int, year: int) -> str:
    return f"data/emp_cache_{_safe_id(sup_id)}_{year}_{month:02d}.csv"

def result_path(sup_id: str) -> str:
    return f"data/final_allocation_{_safe_id(sup_id)}.csv"

def excel_path(sup_id: str) -> str:
    return f"data/Final_Dashboard_{_safe_id(sup_id)}.xlsx"

# ══════════════════════════════════════════════════════════
#  Helpers: Export
# ══════════════════════════════════════════════════════════

def export_result_path(sup_id: str, brand: str) -> str:
    brand_safe = _safe_id(brand) if brand != "ALL" else "ALL"
    return f"data/export_{_safe_id(sup_id)}_{brand_safe}.csv"


# ══════════════════════════════════════════════════════════
#  Helpers: CSV loaders
# ══════════════════════════════════════════════════════════
def load_target_csv() -> tuple[pd.DataFrame | None, pd.DataFrame | None]:
    df_sku = df_sun = None
    if os.path.exists("data/target_boxes.csv"):
        df_sku = (
            pd.read_csv("data/target_boxes.csv", dtype={"sku": str})
            .dropna(subset=["sku"])
            .fillna(0)
        )
        df_sku["sku"] = df_sku["sku"].astype(str).str.strip()
    if os.path.exists("data/target_sun.csv"):
        df_sun = (
            pd.read_csv("data/target_sun.csv", dtype={"emp_id": str})
            .dropna(subset=["emp_id"])
            .fillna(0)
        )
        df_sun["emp_id"] = df_sun["emp_id"].astype(str).str.strip()
    return df_sku, df_sun


def generate_dummy_targets(
    emp_list: list,
    df_sku_base: pd.DataFrame,
    df_hist: pd.DataFrame,
    seed: int = 42,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """สร้าง dummy CSV อัตโนมัติ อิงประวัติจริง ถ้ามี"""
    random.seed(seed)

    records_sku = []
    for _, row in df_sku_base.iterrows():
        sku   = str(row["sku"])
        price = float(row.get("price_per_box") or row.get("unit_cost") or 0)
        if price == 0:
            price = PRICE_FALLBACK.get(sku, round(random.uniform(200, 400), 2))

        hist_total = df_hist[df_hist["sku"] == sku]["hist_boxes"].sum() if not df_hist.empty else 0
        avg_3m     = hist_total / 3.0 if hist_total > 0 else random.randint(10, 100)
        target_boxes = max(5, int(avg_3m * random.uniform(1.05, 1.20)))

        records_sku.append({
            "sku":                     sku,
            "price_per_box":           price,
            "supervisor_target_boxes": target_boxes,
            "brand_name_thai":         str(row.get("brand_name_thai", "")),
            "brand_name_english":      str(row.get("brand_name_english", "")),
            "product_name_thai":       str(row.get("product_name_thai", "")),
        })

    df_sku_out     = pd.DataFrame(records_sku)
    total_box_value = (df_sku_out["price_per_box"] * df_sku_out["supervisor_target_boxes"]).sum()

    # กระจาย target_sun ตามสัดส่วนสุ่ม (supervisor จะปรับเองใน step 2)
    weights = [max(0.5, random.uniform(0.5, 1.5)) for _ in emp_list]
    total_w = sum(weights)
    records_sun, acc = [], 0.0
    for i, (emp, w) in enumerate(zip(emp_list, weights)):
        if i == len(emp_list) - 1:
            ts = round(total_box_value - acc, 2)
        else:
            ts = round(total_box_value * w / total_w, 2)
            acc += ts
        records_sun.append({"emp_id": emp, "target_sun": ts})

    df_sun = pd.DataFrame(records_sun)

    # ตรวจสอบ sum ตรงกัน
    sun_total = df_sun["target_sun"].sum()
    deviation = abs(sun_total - total_box_value)
    if deviation > 1.0:
        logger.warning("target_sun sum deviation: %.2f (expected %.2f)", sun_total, total_box_value)

    os.makedirs("data", exist_ok=True)
    df_sku_out.to_csv("data/target_boxes.csv", index=False)
    df_sun.to_csv("data/target_sun.csv", index=False)
    logger.info("Dummy targets: %d SKUs, %d employees, total %.2f", len(df_sku_out), len(df_sun), total_box_value)
    return df_sku_out, df_sun


# ══════════════════════════════════════════════════════════
#  GET /managers (ดึงรายชื่อ Super Code ทั้งหมดแบบอัตโนมัติ)
# ══════════════════════════════════════════════════════════
#  (moved to top — see _cleanup_old_caches and lifespan above app init)


# ══════════════════════════════════════════════════════════
#  POST /upload/targets — รับไฟล์ Excel จาก Supervisor
# ══════════════════════════════════════════════════════════
# รูปแบบ Excel ที่รองรับ (2 sheet หรือ 2 ไฟล์แยกกัน):
#
#  Sheet "target_boxes":
#    sku | price_per_box | supervisor_target_boxes | brand_name_thai | brand_name_english | product_name_thai
#
#  Sheet "target_sun":
#    emp_id | target_sun

_BOXES_REQUIRED_COLS = {"sku", "price_per_box", "supervisor_target_boxes"}
_SUN_REQUIRED_COLS   = {"emp_id", "target_sun"}


def _validate_and_save_boxes(df: pd.DataFrame) -> dict:
    df.columns = df.columns.str.strip().str.lower()
    missing = _BOXES_REQUIRED_COLS - set(df.columns)
    if missing:
        raise HTTPException(422, detail=f"target_boxes ขาดคอลัมน์: {', '.join(sorted(missing))}")
    df["sku"] = df["sku"].astype(str).str.strip()
    df = df.dropna(subset=["sku"]).copy()
    df["price_per_box"] = pd.to_numeric(df["price_per_box"], errors="coerce").fillna(0)
    df["supervisor_target_boxes"] = pd.to_numeric(df["supervisor_target_boxes"], errors="coerce").fillna(0)
    if (df["price_per_box"] <= 0).any():
        bad = df.loc[df["price_per_box"] <= 0, "sku"].tolist()
        raise HTTPException(422, detail=f"price_per_box ต้อง > 0 — SKU ที่มีปัญหา: {bad[:5]}")
    if (df["supervisor_target_boxes"] < 0).any():
        raise HTTPException(422, detail="supervisor_target_boxes ต้อง >= 0 ทุกแถว")
    for col in ["brand_name_thai", "brand_name_english", "product_name_thai"]:
        if col not in df.columns:
            df[col] = ""
    os.makedirs("data", exist_ok=True)
    df.to_csv("data/target_boxes.csv", index=False)
    total_value = (df["price_per_box"] * df["supervisor_target_boxes"]).sum()
    logger.info("upload target_boxes: %d SKUs, total value=%.2f", len(df), total_value)
    return {"skus": len(df), "total_value": round(total_value, 2)}


def _validate_and_save_sun(df: pd.DataFrame) -> dict:
    df.columns = df.columns.str.strip().str.lower()
    missing = _SUN_REQUIRED_COLS - set(df.columns)
    if missing:
        raise HTTPException(422, detail=f"target_sun ขาดคอลัมน์: {', '.join(sorted(missing))}")
    df["emp_id"] = df["emp_id"].astype(str).str.strip()
    df = df.dropna(subset=["emp_id"]).copy()
    df["target_sun"] = pd.to_numeric(df["target_sun"], errors="coerce").fillna(0)
    if (df["target_sun"] < 0).any():
        raise HTTPException(422, detail="target_sun ต้อง >= 0 ทุกแถว")
    os.makedirs("data", exist_ok=True)
    df.to_csv("data/target_sun.csv", index=False)
    total_sun = df["target_sun"].sum()
    logger.info("upload target_sun: %d employees, total=%.2f", len(df), total_sun)
    return {"employees": len(df), "total_sun": round(total_sun, 2)}


@app.post("/upload/targets")
async def upload_targets(
    file:      UploadFile = File(...),
    file_type: str        = Query(..., description="'boxes' | 'sun' | 'both'"),
    sup_id:    str        = Query(""),
):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, detail="รองรับเฉพาะไฟล์ .xlsx หรือ .xls เท่านั้น")
    if file_type not in ("boxes", "sun", "both"):
        raise HTTPException(400, detail="file_type ต้องเป็น 'boxes', 'sun', หรือ 'both'")
    try:
        contents = await file.read()
        xls = pd.ExcelFile(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(400, detail=f"อ่านไฟล์ Excel ไม่ได้: {e}")

    result = {}
    if file_type in ("boxes", "both"):
        sheet = "target_boxes" if "target_boxes" in xls.sheet_names else xls.sheet_names[0]
        try:
            df_boxes = pd.read_excel(xls, sheet_name=sheet, dtype={"sku": str})
        except Exception as e:
            raise HTTPException(400, detail=f"อ่าน sheet '{sheet}' ไม่ได้: {e}")
        result["boxes"] = _validate_and_save_boxes(df_boxes)

    if file_type in ("sun", "both"):
        sheet = "target_sun" if "target_sun" in xls.sheet_names else (
            xls.sheet_names[1] if len(xls.sheet_names) > 1 else xls.sheet_names[0]
        )
        try:
            df_sun = pd.read_excel(xls, sheet_name=sheet, dtype={"emp_id": str})
        except Exception as e:
            raise HTTPException(400, detail=f"อ่าน sheet '{sheet}' ไม่ได้: {e}")
        result["sun"] = _validate_and_save_sun(df_sun)

    logger.info("upload/targets: sup=%s type=%s result=%s", sup_id, file_type, result)
    return {"status": "ok", "file_type": file_type, **result}


# ══════════════════════════════════════════════════════════
#  GET /upload/template — ดาวน์โหลด Excel template
# ══════════════════════════════════════════════════════════
@app.get("/upload/template")
def download_template():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    def _sheet(ws, headers, required, examples):
        FILL_REQ = PatternFill("solid", fgColor="2E75B6")
        FILL_OPT = PatternFill("solid", fgColor="1F4E79")
        FILL_EX  = PatternFill("solid", fgColor="EBF3FB")
        FNT      = Font(color="FFFFFF", bold=True, size=11)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = FNT; c.fill = FILL_REQ if h in required else FILL_OPT
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(ci)].width = max(20, len(h) + 4)
        ws.row_dimensions[1].height = 28
        for ri, row in enumerate(examples, 2):
            for ci, v in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=v).fill = FILL_EX

    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "target_boxes"
    _sheet(ws1,
           ["sku","price_per_box","supervisor_target_boxes","brand_name_thai","brand_name_english","product_name_thai"],
           {"sku","price_per_box","supervisor_target_boxes"},
           [["624007", 240.00, 150, "แบรนด์ A", "Brand A", "สินค้า A"],
            ["624015", 212.00, 80,  "แบรนด์ B", "Brand B", "สินค้า B"]])

    ws2 = wb.create_sheet("target_sun")
    _sheet(ws2, ["emp_id","target_sun"], {"emp_id","target_sun"},
           [["EMP001", 125000.00], ["EMP002", 98000.00]])

    os.makedirs("data", exist_ok=True)
    tpl_path = "data/upload_template.xlsx"
    wb.save(tpl_path)
    return FileResponse(tpl_path, filename="Target_Upload_Template.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ══════════════════════════════════════════════════════════
#  GET /managers (ดึงรายชื่อ Super Code ทั้งหมดแบบอัตโนมัติ)
# ══════════════════════════════════════════════════════════

@app.get("/managers")
def get_managers():
    os.makedirs("data", exist_ok=True)
    cache_path = "data/managers_cache.json"
    
    try:
        fabric = FabricDAXConnector()
        managers = fabric.get_all_super_codes()
        
        if managers:
            # เซฟเก็บไว้ในไฟล์ เผื่อรอบหน้า Power BI ตอบช้า
            with open(cache_path, "w", encoding="utf-8") as f:
                json.dump(managers, f)
            return {"managers": managers}
            
    except Exception as e:
        logger.warning("get_all_super_codes error: %s", e)
        
    # ถ้ายิง Fabric ไม่สำเร็จ ให้โหลดจาก Cache ล่าสุดมาใช้แทน
    if os.path.exists(cache_path):
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                return {"managers": json.load(f)}
        except Exception as cache_err:
            logger.warning("managers cache corrupt: %s", cache_err)
            
    # ถ้าพังหมดเลย ให้คืนค่า Default กันหน้าเว็บพัง
    return {"managers": KNOWN_MANAGERS}


# ══════════════════════════════════════════════════════════
#  GET /data/employees
# ══════════════════════════════════════════════════════════
@app.get("/data/employees")
def get_employees(
    sup_id:       str  = Query(..., description="SuperCode เช่น SL330"),
    target_month: int  = Query(..., ge=1, le=12),
    target_year:  int  = Query(..., ge=2020, le=2100),
    regen_target: bool = Query(False, description="บังคับ regenerate dummy targets"),
):
    # ── Validate supervisor ───────────────────────────────
    # ใช้ warning แทน error เพราะ SuperCode อาจถูกเพิ่มในภายหลัง
    if sup_id not in KNOWN_MANAGERS:
        logger.warning("sup_id '%s' ไม่อยู่ใน KNOWN_MANAGERS %s → อนุญาตให้ดำเนินการต่อ", sup_id, KNOWN_MANAGERS)

    os.makedirs("data", exist_ok=True)

    # ── Step 1: ดึงพนักงาน ───────────────────────────────
    fabric = None
    df_emp_fabric = pd.DataFrame()
    try:
        fabric = FabricDAXConnector()
        df_emp_fabric = fabric.get_employees_by_manager(sup_id)
    except Exception as e:
        cp = emp_cache_path(sup_id, target_month, target_year)
        if os.path.exists(cp):
            logger.warning("Fabric error → emp cache: %s", e)
            df_emp_fabric = pd.read_csv(cp, dtype={"emp_id": str})
        else:
            raise HTTPException(503, detail=f"ไม่สามารถดึงพนักงานได้ และไม่มี cache: {e}")

    if df_emp_fabric.empty:
        raise HTTPException(404, detail=f"ไม่พบพนักงานใต้ SuperCode '{sup_id}'")

    emp_list = df_emp_fabric["emp_id"].tolist()
    df_emp_fabric.to_csv(emp_cache_path(sup_id, target_month, target_year), index=False)
    logger.info("Employees: %d คน %s", len(emp_list), emp_list)

    # ── Step 2: โหลด/สร้าง SKU targets ──────────────────
    df_sku_csv, df_sun_csv = load_target_csv()

    if df_sku_csv is not None and not regen_target:
        logger.info("ใช้ target_boxes.csv ที่มีอยู่")
        df_sku  = df_sku_csv
        sku_list = df_sku["sku"].tolist()
    else:
        # ดึง SKU จาก Fabric
        try:
            sku_list = fabric.get_skus_sold_by_team(emp_list, target_month, target_year, n_months=6)
        except Exception as e:
            logger.warning("get_skus_sold_by_team error: %s → fallback SKUs", e)
            sku_list = list(PRICE_FALLBACK.keys())

        if not sku_list:
            sku_list = list(PRICE_FALLBACK.keys())

        # ดึง product info
        df_sku_base = pd.DataFrame()
        try:
            df_sku_base = fabric.get_product_info(sku_list=sku_list)
        except Exception as e:
            logger.warning("get_product_info error: %s", e)
            df_sku_base = pd.DataFrame({"sku": sku_list})

        if df_sku_base.empty:
            df_sku_base = pd.DataFrame({"sku": sku_list})

        # ดึงประวัติก่อน generate dummy targets
        df_hist_pre = pd.DataFrame()
        try:
            df_hist_pre = fabric.get_historical_sales(
                target_month, target_year, sku_list=sku_list, emp_list=emp_list
            )
        except Exception as e:
            logger.warning("historical (pre-gen): %s", e)

        df_sku, df_sun_csv = generate_dummy_targets(emp_list, df_sku_base, df_hist_pre)

    # refresh sun CSV
    if df_sun_csv is None and os.path.exists("data/target_sun.csv"):
        df_sun_csv = pd.read_csv("data/target_sun.csv", dtype={"emp_id": str}).fillna(0)
        df_sun_csv["emp_id"] = df_sun_csv["emp_id"].astype(str).str.strip()

    sku_list = df_sku["sku"].tolist()

    # ── Step 3: merge target_sun ──────────────────────────
    df_emp = df_emp_fabric.copy()
    if df_sun_csv is not None and not df_sun_csv.empty:
        df_emp = pd.merge(df_emp, df_sun_csv[["emp_id", "target_sun"]], on="emp_id", how="left")
    if "target_sun" not in df_emp.columns:
        df_emp["target_sun"] = 0.0
    df_emp["target_sun"] = df_emp["target_sun"].fillna(0.0)

    # ── Step 4: LY sales ──────────────────────────────────
    # get_ly_sales แล้ว fill 0 ให้ครบทุก emp (handled ใน connector)
    try:
        df_ly = fabric.get_ly_sales(target_month, target_year, sku_list=sku_list, emp_list=emp_list)
        if not df_ly.empty and "emp_id" in df_ly.columns:
            # merge เฉพาะ emp_id + ly_sales (connector ทำ fill 0 ให้แล้ว)
            df_emp = pd.merge(df_emp, df_ly[["emp_id", "ly_sales"]], on="emp_id", how="left")
        else:
            df_emp["ly_sales"] = 0.0
    except Exception as e:
        logger.warning("LY sales error: %s", e)
        df_emp["ly_sales"] = 0.0
    df_emp["ly_sales"] = df_emp["ly_sales"].fillna(0.0)

    # ── Step 5: Historical 3M ─────────────────────────────
    df_hist = pd.DataFrame()
    try:
        df_hist = fabric.get_historical_sales(
            target_month, target_year, sku_list=sku_list, emp_list=emp_list
        )
        if not df_hist.empty:
            df_hist_emp = df_hist[df_hist["emp_id"].isin(emp_list)].copy()
            df_hist_emp.to_csv(hist_cache_path(sup_id, target_month, target_year), index=False)

            df_hist_val = pd.merge(df_hist_emp, df_sku[["sku", "price_per_box"]], on="sku", how="left")
            df_hist_val["hist_value"] = df_hist_val["hist_boxes"] * df_hist_val["price_per_box"]
            df_hist_agg = df_hist_val.groupby("emp_id")["hist_value"].sum().reset_index()
            df_hist_agg["hist_avg_3m"] = df_hist_agg["hist_value"] / 3.0
            df_emp = pd.merge(df_emp, df_hist_agg[["emp_id", "hist_avg_3m"]], on="emp_id", how="left")
        else:
            df_emp["hist_avg_3m"] = 0.0
    except Exception as e:
        logger.warning("historical 3M: %s", e)
        df_emp["hist_avg_3m"] = 0.0
    df_emp["hist_avg_3m"] = df_emp["hist_avg_3m"].fillna(0.0)

    # ── Step 6: Warehouse ─────────────────────────────────
    try:
        df_wh = fabric.get_warehouse_by_emp(emp_list)
        if not df_wh.empty:
            df_emp = pd.merge(df_emp, df_wh[["emp_id", "warehouse_code"]], on="emp_id", how="left")
    except Exception as e:
        logger.warning("warehouse: %s", e)
    if "warehouse_code" not in df_emp.columns:
        df_emp["warehouse_code"] = ""
    df_emp["warehouse_code"] = df_emp["warehouse_code"].fillna("")

    # ── Cleanup dtypes ────────────────────────────────────
    numeric_cols = df_emp.select_dtypes(include=["number"]).columns
    df_emp[numeric_cols] = df_emp[numeric_cols].fillna(0)
    for col in ["emp_name", "manager_code", "warehouse_code"]:
        if col in df_emp.columns:
            df_emp[col] = df_emp[col].fillna("")

    logger.info("Response: %d emp, %d sku", len(df_emp), len(df_sku))

    # ── Fix 3: เช็ค emp_id ใน target_sun.csv ที่ไม่ตรงกับพนักงานจริง ──
    sku_warnings = []
    if df_sun_csv is not None and not df_sun_csv.empty:
        sun_emp_ids  = set(df_sun_csv["emp_id"].astype(str).str.strip())
        fabric_emp_ids = set(str(e) for e in emp_list)
        unmatched = sun_emp_ids - fabric_emp_ids
        if unmatched:
            logger.warning("target_sun emp_id ไม่ตรงกับ Fabric: %s", unmatched)
            sku_warnings.append({
                "type": "emp_mismatch",
                "sku": "",
                "brand": "",
                "message": (
                    f"⚠️ target_sun.csv มีรหัสพนักงานที่ไม่มีในทีม {sup_id}: "
                    f"{', '.join(sorted(unmatched)[:5])} "
                    f"— เป้าเงินของพนักงานเหล่านี้จะถูกตั้งเป็น 0"
                )
            })

    # ── Fix 4: SKU Reconciliation — รันทุกครั้งที่มี target_boxes.csv ──
    # (ไม่รอให้ hist ไม่ว่าง เพราะถ้า Fabric ล่มก็ยังควรเตือนได้)
    target_skus = set(str(s) for s in df_sku["sku"].tolist())

    if not df_hist.empty:
        # ประวัติจาก Fabric — เช็คทีมทั้งหมด (ไม่ใช่รายคน)
        hist_skus = set(df_hist["sku"].dropna().astype(str).unique())

        # SKU มีเป้าแต่ไม่มีประวัติขายในทีมนี้เลย
        new_skus = sorted(target_skus - hist_skus)
        for sku in new_skus:
            info = df_sku[df_sku["sku"] == sku]
            brand = ""
            if not info.empty:
                brand = str(info.iloc[0].get("brand_name_thai") or info.iloc[0].get("brand_name_english") or "")
            sku_warnings.append({
                "type": "no_history",
                "sku": sku,
                "brand": brand,
                "message": f"SKU {sku}{' (' + brand + ')' if brand else ''} มีเป้าหีบ แต่ไม่มีพนักงานในทีมนี้เคยขายมาก่อนเลย"
            })

        # SKU เคยขายในทีมแต่ไม่มีในเป้าเดือนนี้
        missing_skus = sorted(hist_skus - target_skus)
        for sku in missing_skus:
            hist_rows = df_hist[df_hist["sku"] == sku]
            total_hist = float(hist_rows["hist_boxes"].sum()) if "hist_boxes" in hist_rows.columns else 0
            sku_warnings.append({
                "type": "no_target",
                "sku": sku,
                "brand": "",
                "message": f"SKU {sku} เคยขายได้ {total_hist:.0f} หีบ (3M) แต่ไม่ถูกรวมในเป้าเดือนนี้"
            })
    else:
        # Fabric ดึงประวัติไม่ได้ — เตือนเฉพาะกรณีไม่มี hist cache เลย
        cache_file = hist_cache_path(sup_id, target_month, target_year)
        if not os.path.exists(cache_file) and target_skus:
            sku_warnings.append({
                "type": "no_history",
                "sku": "",
                "brand": "",
                "message": "⚠️ ไม่สามารถดึงประวัติขายจาก Fabric ได้ — การกระจายหีบจะใช้ EVEN แทนประวัติ"
            })

    if sku_warnings:
        logger.info("reconciliation warnings: %d รายการ", len(sku_warnings))

    def _clean(df: pd.DataFrame) -> list:
        """แปลง NaN → None ก่อน serialize เพื่อกัน JSON invalid"""
        return df.where(pd.notna(df), None).to_dict(orient="records")

    return {
        "employees":    _clean(df_emp),
        "skus":         _clean(df_sku),
        "sku_warnings": sku_warnings,
    }


# ══════════════════════════════════════════════════════════
#  POST /optimize
# ══════════════════════════════════════════════════════════
@app.post("/optimize")
def run_optimization(
    req:          OptimizeRequest,
    sup_id:       str = Query("SL330"),
    target_month: int = Query(..., ge=1, le=12),
    target_year:  int = Query(..., ge=2020, le=2100),
):
    # Validate
    if sup_id not in KNOWN_MANAGERS:
        logger.warning("sup_id '%s' ไม่อยู่ใน KNOWN_MANAGERS → อนุญาตให้ดำเนินการต่อ", sup_id)
    if req.strategy.upper() not in VALID_STRATEGIES:
        raise HTTPException(400, detail=f"strategy ไม่ถูกต้อง ต้องเป็น {VALID_STRATEGIES}")

    os.makedirs("data", exist_ok=True)

    df_sku, _ = load_target_csv()
    if df_sku is None:
        raise HTTPException(500, detail="ไม่พบ target_boxes.csv กรุณาโหลดหน้า Dashboard ก่อน")

    df_emp_targets = pd.DataFrame([t.dict() for t in req.yellowTargets])
    emp_list = df_emp_targets["emp_id"].tolist()

    # โหลด historical cache (ต่อ supervisor)
    cache_file = hist_cache_path(sup_id, target_month, target_year)
    if os.path.exists(cache_file):
        df_hist = pd.read_csv(cache_file, dtype={"sku": str, "emp_id": str})
        logger.info("hist cache loaded: %d rows", len(df_hist))
    else:
        logger.warning("ไม่พบ hist cache → ใช้ตารางเปล่า")
        df_hist = pd.DataFrame(columns=["emp_id", "sku", "hist_boxes"])

    df_hist = df_hist[df_hist["emp_id"].isin(emp_list)]

    # รัน allocation ตาม strategy ที่เลือก
    logger.info("Running strategy=%s for sup=%s", req.strategy, sup_id)
    locked_edits_data = [{"emp_id": le.emp_id, "sku": le.sku, "locked_boxes": le.locked_boxes} for le in req.locked_edits]
    df_allocation = allocate_boxes(
        df_emp_targets, df_sku, df_hist,
        strategy=req.strategy,
        force_min_one=req.force_min_one,
        locked_edits=locked_edits_data if locked_edits_data else None,
        cap_multiplier=req.cap_multiplier,
    )

    # คำนวณ hist_avg 3M รายพนักงาน×SKU
    if not df_hist.empty:
        df_hist_avg = df_hist.groupby(["emp_id", "sku"])["hist_boxes"].sum().reset_index()
        df_hist_avg["hist_avg"] = (df_hist_avg["hist_boxes"] / 3.0).round(1)
    else:
        df_hist_avg = pd.DataFrame(columns=["emp_id", "sku", "hist_avg"])

    df_final = pd.merge(
        df_allocation,
        df_hist_avg[["emp_id", "sku", "hist_avg"]],
        on=["emp_id", "sku"], how="left"
    )
    df_final["hist_avg"] = df_final["hist_avg"].fillna(0)

    # merge brand + price_per_box
    brand_cols = [c for c in ["brand_name_thai", "brand_name_english", "product_name_thai", "price_per_box"]
                  if c in df_sku.columns]
    if brand_cols:
        df_final = pd.merge(df_final, df_sku[["sku"] + brand_cols], on="sku", how="left")
        for c in brand_cols:
            df_final[c] = df_final[c].fillna("" if "name" in c else 0)

    # บันทึก per-supervisor (ป้องกัน race condition)
    df_final.to_csv(result_path(sup_id), index=False)

    # สร้าง Excel ทั้งหมด (ALL brand) ทันที
    yellow_map = {y.emp_id: y.yellow_target for y in req.yellowTargets}
    create_target_excel(
        result_csv=result_path(sup_id),
        output_path=excel_path(sup_id),
        brand_filter="ALL",
        yellow_map=yellow_map,
        sup_id=sup_id,
    )

    return {"allocations": df_final.to_dict(orient="records")}


# ══════════════════════════════════════════════════════════
#  POST /export/excel
# ══════════════════════════════════════════════════════════
@app.post("/export/excel")
def export_excel(
    req:    ExportRequest,
    sup_id: str = Query("SL330"),
):
    if sup_id not in KNOWN_MANAGERS:
        logger.warning("export: sup_id '%s' ไม่อยู่ใน KNOWN_MANAGERS → อนุญาตให้ดำเนินการต่อ", sup_id)

    os.makedirs("data", exist_ok=True)

    # Build DataFrame จาก validated Pydantic models
    df_final = pd.DataFrame([a.dict() for a in req.allocations])

    # เติม price_per_box ถ้าไม่ครบ
    df_sku_tmp, _ = load_target_csv()
    if df_sku_tmp is not None:
        missing_cols = [c for c in ["price_per_box", "brand_name_thai", "brand_name_english", "product_name_thai"]
                        if c not in df_final.columns or (df_final[c] == 0).all()]
        if missing_cols:
            merge_cols = ["sku"] + [c for c in missing_cols if c in df_sku_tmp.columns]
            df_final = pd.merge(df_final, df_sku_tmp[merge_cols], on="sku", how="left", suffixes=("", "_csv"))
            for c in missing_cols:
                if f"{c}_csv" in df_final.columns:
                    df_final[c] = df_final[c].where(df_final[c] != 0, df_final[f"{c}_csv"])
                    df_final.drop(columns=[f"{c}_csv"], inplace=True)

    # filter by brand
    brand_filter = req.brand_filter
    df_export = df_final.copy()
    if brand_filter != "ALL":
        df_export = df_final[df_final["brand_name_thai"] == brand_filter].copy()
        if df_export.empty:
            raise HTTPException(404, detail=f"ไม่พบข้อมูลสำหรับแบรนด์ '{brand_filter}'")

    ep = export_result_path(sup_id, brand_filter)
    df_export.to_csv(ep, index=False)

    yellow_map = {y.emp_id: y.yellow_target for y in req.yellow_targets}

    create_target_excel(
        result_csv=ep,
        output_path=excel_path(sup_id),
        brand_filter=brand_filter,
        yellow_map=yellow_map,
        sup_id=sup_id,
    )

    logger.info("Export excel: sup=%s brand=%s rows=%d", sup_id, brand_filter, len(df_export))
    return {"status": "ok", "brand_filter": brand_filter, "rows": len(df_export)}


# ══════════════════════════════════════════════════════════
#  GET /download/excel
# ══════════════════════════════════════════════════════════
@app.get("/download/excel")
def download_excel(sup_id: str = Query("SL330")):
    if sup_id not in KNOWN_MANAGERS:
        logger.warning("download: sup_id '%s' ไม่อยู่ใน KNOWN_MANAGERS → อนุญาตให้ดำเนินการต่อ", sup_id)

    fpath = excel_path(sup_id)
    if not os.path.exists(fpath):
        raise HTTPException(404, detail="ไม่พบไฟล์ Excel กรุณา Optimize ก่อน")

    return FileResponse(
        fpath,
        filename=f"Target_{_safe_id(sup_id)}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ══════════════════════════════════════════════════════════
#  GET /health
# ══════════════════════════════════════════════════════════
@app.get("/health")
def health():
    return {
        "status": "ok",
        "known_managers": KNOWN_MANAGERS,
        "valid_strategies": list(VALID_STRATEGIES),
        "files": {
            "target_boxes.csv": os.path.exists("data/target_boxes.csv"),
            "target_sun.csv":   os.path.exists("data/target_sun.csv"),
        },
    }


# ══════════════════════════════════════════════════════════
#  GET /debug/fabric  — ดู SuperCode ทั้งหมดที่มีใน dim_salesman
#  ใช้สำหรับ debug เมื่อหาพนักงานไม่เจอ
# ══════════════════════════════════════════════════════════
@app.get("/debug/fabric")
def debug_fabric(sup_id: str = Query("SL330")):
    """
    ดึงข้อมูล debug จาก Fabric:
    - SuperCode ทั้งหมดที่มีใน dim_salesman
    - พนักงานที่ SuperCode นี้ดูแล (ถ้าเจอ)
    เปิด: http://localhost:8000/debug/fabric?sup_id=SL330
    """
    try:
        fabric = FabricDAXConnector()

        # ดึง SuperCode sample ทั้งหมด
        dax_all = """
EVALUATE
SUMMARIZECOLUMNS(
    'dim_salesman'[SuperCode],
    "cnt", COUNTROWS('dim_salesman')
)
ORDER BY 'dim_salesman'[SuperCode]
"""
        rows_all = fabric._execute_dax(dax_all, debug=True)
        all_super_codes = []
        for r in rows_all:
            sc = fabric._get(r, "[SuperCode]", "dim_salesman[SuperCode]", default="")
            cnt = fabric._get(r, "[cnt]", default=0)
            all_super_codes.append({"super_code": repr(str(sc)), "count": cnt})

        # ลอง query พนักงาน
        df_emp = fabric.get_employees_by_manager(sup_id)

        return {
            "query_super_code": sup_id,
            "query_super_code_repr": repr(sup_id),
            "employees_found": len(df_emp),
            "employees": df_emp.to_dict(orient="records") if not df_emp.empty else [],
            "all_super_codes_in_db": all_super_codes,
        }
    except Exception as e:
        logger.error("debug_fabric error: %s", e)
        raise HTTPException(500, detail=str(e))


# ══════════════════════════════════════════════════════════
#  สำหรับรันเป็นไฟล์ .exe โดยไม่ต้องใช้คำสั่งผ่าน Terminal
# ══════════════════════════════════════════════════════════
if __name__ == "__main__":
    import uvicorn
    # เปิดเซิร์ฟเวอร์ที่ port 8000 (เอา --reload ออกเพราะเป็น production)
    uvicorn.run(app, host="127.0.0.1", port=8000)