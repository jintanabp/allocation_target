"""
generate_excel.py — สร้าง Excel กระจายเป้าหมาย
────────────────────────────────────────────────
ตาม template รูปที่ 2:
  แถว 1   : Title + supervisor info
  แถว 4   : ราคา/หีบ
  แถว 5   : รหัส SKU (bold, สีเหลือง)
  แถว 7   : เป้ารวม (supervisor_target_boxes)
  แถว 8   : เป้าย่อย (sum จากพนักงาน)
  แถว 9   : เคยขาย (hist_avg ทีม)
  แถว 10+ : พนักงาน (คู่: เป้า / เคยขาย)
  คอลัมน์ท้าย : ยอดรวม (บาท)
"""

import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ── Styles ───────────────────────────────────────────────
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
GREEN_FILL  = PatternFill("solid", fgColor="92D050")
BLUE_FILL   = PatternFill("solid", fgColor="BDD7EE")
GREY_FILL   = PatternFill("solid", fgColor="D9D9D9")
ORANGE_FILL = PatternFill("solid", fgColor="F4B942")
HDR_FILL    = PatternFill("solid", fgColor="1F497D")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
WARN_FILL   = PatternFill("solid", fgColor="FFEB9C")  # warning: ยอดเงินห่างจากเป้า

THIN = Side(style="thin", color="BFBFBF")
MED  = Side(style="medium", color="595959")
THIN_BRD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED_BRD  = Border(left=MED,  right=MED,  top=MED,  bottom=MED)

BOLD_HDR  = Font(name="Cordia New", bold=True, color="FFFFFF", size=12)
BOLD_BLK  = Font(name="Cordia New", bold=True, size=11)
NORM      = Font(name="Cordia New", size=11)
SMALL     = Font(name="Cordia New", size=10, italic=True, color="595959")
NUM_FMT   = "#,##0"
NUM_FMT1  = "#,##0.0"
BAHT_FMT  = "#,##0.00"

CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left",   vertical="center")
RGT = Alignment(horizontal="right",  vertical="center")


def _c(ws, row, col, value=None, font=None, fill=None, align=None, border=None, num_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:    cell.font      = font
    if fill:    cell.fill      = fill
    if align:   cell.alignment = align
    if border:  cell.border    = border
    if num_fmt: cell.number_format = num_fmt
    return cell


# ══════════════════════════════════════════════════════════
#  PUBLIC ENTRY POINT
# ══════════════════════════════════════════════════════════
def create_target_excel(
    result_csv:  str,
    output_path: str,
    brand_filter: str = "ALL",
    yellow_map:  dict | None = None,  # emp_id → yellow_target
    sup_id:      str = "",
    target_boxes_csv: str | None = "data/target_boxes.csv",
) -> str | None:
    """
    สร้างไฟล์ Excel จาก result_csv
    brand_filter = "ALL" → ทุกแบรนด์ | "ชื่อแบรนด์" → กรองเฉพาะแบรนด์นั้น
    yellow_map   = {emp_id: yellow_target_baht} สำหรับแสดงใน header และ validate deviation
    target_boxes_csv — ถ้ามีไฟล์ จะอ่าน supervisor_target_boxes มาแสดงเป็นแถว "เป้าหีบ (หัวหน้า)"
    แถวถัดไปเป็น "ผลรวมกระจาย" เพื่อให้เทียบกับผลคำนวณได้
    """
    if not os.path.exists(result_csv):
        print(f"❌ ไม่พบ {result_csv}")
        return None

    df = pd.read_csv(result_csv, dtype={"sku": str, "emp_id": str})
    df["allocated_boxes"] = pd.to_numeric(df["allocated_boxes"], errors="coerce").fillna(0).astype(int)
    df["hist_avg"]        = pd.to_numeric(df["hist_avg"],        errors="coerce").fillna(0.0)
    df["price_per_box"]   = pd.to_numeric(df["price_per_box"],   errors="coerce").fillna(0.0)
    df["brand_name_thai"] = df.get("brand_name_thai", pd.Series("", index=df.index)).fillna("").astype(str)

    # ── กรองแบรนด์ ────────────────────────────────────────
    if brand_filter != "ALL":
        df = df[df["brand_name_thai"] == brand_filter].copy()
        if df.empty:
            print(f"⚠️ ไม่พบข้อมูลสำหรับแบรนด์ '{brand_filter}'")
            return None

    if df.empty:
        print("⚠️ ไม่มีข้อมูล allocation")
        return None

    skus = df["sku"].unique().tolist()
    emps = df["emp_id"].unique().tolist()
    sku_price  = dict(zip(df["sku"], df["price_per_box"]))
    sku_brand  = dict(zip(df["sku"], df["brand_name_thai"]))
    yellow_map = yellow_map or {}

    # ผลรวมหีบที่กระจายแล้วต่อ SKU
    sku_allocated = df.groupby("sku")["allocated_boxes"].sum().to_dict()

    # เป้าหีบจากหัวหน้า (จาก target_boxes.csv) — ถ้าไม่มีไฟล์จะใช้ผลรวมกระจายแทน
    sku_official: dict[str, int] = {}
    tpath = target_boxes_csv or ""
    if tpath and os.path.exists(tpath):
        try:
            df_t = pd.read_csv(tpath, dtype={"sku": str}).dropna(subset=["sku"])
            df_t["sku"] = df_t["sku"].astype(str).str.strip()
            if brand_filter != "ALL" and "brand_name_thai" in df_t.columns:
                df_t = df_t[df_t["brand_name_thai"].astype(str) == str(brand_filter)]
            for _, r in df_t.iterrows():
                sku_official[str(r["sku"]).strip()] = int(round(float(r.get("supervisor_target_boxes", 0) or 0)))
        except Exception as ex:
            print(f"⚠️ อ่าน {tpath} สำหรับเป้าหีบหัวหน้าไม่ได้: {ex}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "กระจายเป้า"

    # ── Column widths ────────────────────────────────────
    ws.column_dimensions["A"].width = 5   # ลำดับ
    ws.column_dimensions["B"].width = 10  # S/M
    ws.column_dimensions["C"].width = 10  # กลุ่มย่อย
    ws.column_dimensions["D"].width = 10  # W/H
    ws.column_dimensions["E"].width = 8   # ประเภท

    DATA_COL_START = 6  # F
    for i in range(len(skus)):
        ws.column_dimensions[get_column_letter(DATA_COL_START + i)].width = 10
    total_col = DATA_COL_START + len(skus)
    ws.column_dimensions[get_column_letter(total_col)].width = 18

    # ── แถว 1: Title ────────────────────────────────────
    ws.row_dimensions[1].height = 28
    brand_label = brand_filter if brand_filter != "ALL" else "ทุกแบรนด์"
    total_yellow = sum(yellow_map.values())
    title_text = (
        f"ปรับปรุงเป้าหมายพนักงานขายเดือนหน้า  |  Supervisor: {sup_id}  |  "
        f"แบรนด์: {brand_label}  |  เป้ารวม: {total_yellow:,.0f} บาท"
    )
    t = ws.cell(row=1, column=1, value=title_text)
    t.font = Font(name="Cordia New", bold=True, size=13, color="1F497D")
    t.alignment = LFT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_col)

    # ── แถว 2: ว่าง (spacer) ────────────────────────────
    ws.row_dimensions[2].height = 6

    # ── แถว 3: Header A-E + SKU group header ───────────
    ws.row_dimensions[3].height = 6

    # ── แถว 4: ราคา/หีบ ─────────────────────────────────
    ws.row_dimensions[4].height = 20
    for ci, lbl in enumerate(["", "ราคา/หีบ", "", "", ""], start=1):
        fill = GREY_FILL if lbl else WHITE_FILL
        _c(ws, 4, ci, lbl or None, font=BOLD_BLK, fill=fill, align=CTR, border=THIN_BRD)
    for i, sku in enumerate(skus):
        _c(ws, 4, DATA_COL_START + i, sku_price.get(sku, 0),
           font=BOLD_BLK, fill=YELLOW_FILL, align=CTR, border=THIN_BRD, num_fmt=NUM_FMT1)
    _c(ws, 4, total_col, None, fill=WHITE_FILL, border=THIN_BRD)

    # ── แถว 5: รหัสสินค้า SKU ───────────────────────────
    ws.row_dimensions[5].height = 22
    for ci, lbl in enumerate(["ลำดับ", "S/M", "กลุ่มย่อย", "W/H", "ประเภท"], start=1):
        _c(ws, 5, ci, lbl, font=BOLD_BLK, fill=GREY_FILL, align=CTR, border=THIN_BRD)
    for i, sku in enumerate(skus):
        _c(ws, 5, DATA_COL_START + i, sku,
           font=Font(name="Cordia New", bold=True, size=11),
           fill=YELLOW_FILL, align=CTR, border=THIN_BRD)
    _c(ws, 5, total_col, "ยอดรวม (บาท)",
       font=BOLD_HDR, fill=HDR_FILL, align=CTR, border=THIN_BRD)

    # ── แถว 6: ว่าง ─────────────────────────────────────
    ws.row_dimensions[6].height = 4

    # ── แถว 7: เป้ารวม (supervisor total target boxes) ──
    ws.row_dimensions[7].height = 22
    _c(ws, 7, 1, None, fill=YELLOW_FILL, border=THIN_BRD)
    _c(ws, 7, 2, None, fill=YELLOW_FILL, border=THIN_BRD)
    _c(ws, 7, 3, None, fill=YELLOW_FILL, border=THIN_BRD)
    _c(ws, 7, 4, None, fill=YELLOW_FILL, border=THIN_BRD)
    _c(ws, 7, 5, "เป้าหีบ (หัวหน้า)", font=BOLD_BLK, fill=ORANGE_FILL, align=CTR, border=THIN_BRD)
    row7_val = 0
    for i, sku in enumerate(skus):
        boxes = int(sku_official[sku]) if sku in sku_official else int(sku_allocated.get(sku, 0))
        row7_val += boxes * sku_price.get(sku, 0)
        _c(ws, 7, DATA_COL_START + i, boxes,
           font=BOLD_BLK, fill=YELLOW_FILL, align=CTR, border=THIN_BRD, num_fmt=NUM_FMT)
    _c(ws, 7, total_col, row7_val,
       font=BOLD_BLK, fill=YELLOW_FILL, align=RGT, border=THIN_BRD, num_fmt=BAHT_FMT)

    # ── แถว 8: ผลรวมกระจาย (หลังคำนวณ) ────────────
    ws.row_dimensions[8].height = 20
    _c(ws, 8, 5, "ผลรวมกระจาย", font=BOLD_BLK, fill=YELLOW_FILL, align=CTR, border=THIN_BRD)
    for ci in [1,2,3,4]:
        _c(ws, 8, ci, None, fill=YELLOW_FILL, border=THIN_BRD)
    row8_val = 0
    for i, sku in enumerate(skus):
        boxes = int(sku_allocated.get(sku, 0))
        row8_val += boxes * sku_price.get(sku, 0)
        _c(ws, 8, DATA_COL_START + i, boxes,
           font=NORM, fill=YELLOW_FILL, align=CTR, border=THIN_BRD, num_fmt=NUM_FMT)
    _c(ws, 8, total_col, row8_val,
       font=BOLD_BLK, fill=YELLOW_FILL, align=RGT, border=THIN_BRD, num_fmt=BAHT_FMT)

    # ── แถว 9: เคยขาย (hist avg ทีม รวม) ───────────────
    ws.row_dimensions[9].height = 18
    _c(ws, 9, 5, "เคยขาย", font=SMALL, fill=GREY_FILL, align=CTR, border=THIN_BRD)
    for ci in [1,2,3,4]:
        _c(ws, 9, ci, None, fill=GREY_FILL, border=THIN_BRD)
    for i, sku in enumerate(skus):
        team_hist = df[df["sku"] == sku]["hist_avg"].sum()
        _c(ws, 9, DATA_COL_START + i, round(team_hist, 1) if team_hist > 0 else None,
           font=SMALL, fill=GREY_FILL, align=CTR, border=THIN_BRD, num_fmt=NUM_FMT1)
    _c(ws, 9, total_col, None, fill=GREY_FILL, border=THIN_BRD)

    # ── แถว 10+: รายพนักงาน ─────────────────────────────
    current_row = 10
    for idx, emp in enumerate(emps, start=1):
        ws.row_dimensions[current_row].height     = 20
        ws.row_dimensions[current_row + 1].height = 18

        emp_df = df[df["emp_id"] == emp]
        yellow_target = yellow_map.get(emp, 0)

        # คำนวณยอดรวมจริงของพนักงาน (ทุก SKU ใน view นี้)
        emp_value = sum(
            int(emp_df[emp_df["sku"] == sku]["allocated_boxes"].sum()) * sku_price.get(sku, 0)
            for sku in skus
        )
        # deviation จาก yellow_target → เลือก fill
        deviation = abs(emp_value - yellow_target)
        val_fill = GREEN_FILL if deviation <= 1000 else WARN_FILL

        # บรรทัด "เป้า"
        _c(ws, current_row, 1, idx, font=NORM, fill=BLUE_FILL, align=CTR, border=THIN_BRD)
        _c(ws, current_row, 2, emp, font=BOLD_BLK, fill=BLUE_FILL, align=CTR, border=THIN_BRD)
        _c(ws, current_row, 3, None, fill=BLUE_FILL, border=THIN_BRD)
        _c(ws, current_row, 4, None, fill=BLUE_FILL, border=THIN_BRD)
        _c(ws, current_row, 5, "เป้า", font=BOLD_BLK, fill=GREEN_FILL, align=CTR, border=THIN_BRD)
        for i, sku in enumerate(skus):
            row_emp = emp_df[emp_df["sku"] == sku]
            boxes = int(row_emp["allocated_boxes"].sum()) if not row_emp.empty else 0
            cell_fill = GREEN_FILL if boxes > 0 else WHITE_FILL
            _c(ws, current_row, DATA_COL_START + i,
               boxes if boxes > 0 else None,
               font=NORM, fill=cell_fill, align=CTR, border=THIN_BRD, num_fmt=NUM_FMT)
        _c(ws, current_row, total_col, emp_value,
           font=BOLD_BLK, fill=val_fill, align=RGT, border=THIN_BRD, num_fmt=BAHT_FMT)

        # บรรทัด "เคยขาย"
        _c(ws, current_row + 1, 1, None, fill=WHITE_FILL, border=THIN_BRD)
        _c(ws, current_row + 1, 2, emp, font=SMALL, fill=WHITE_FILL, align=CTR, border=THIN_BRD)
        _c(ws, current_row + 1, 3, None, fill=WHITE_FILL, border=THIN_BRD)
        _c(ws, current_row + 1, 4, None, fill=WHITE_FILL, border=THIN_BRD)
        _c(ws, current_row + 1, 5, "เคยขาย", font=SMALL, fill=GREY_FILL, align=CTR, border=THIN_BRD)
        for i, sku in enumerate(skus):
            row_emp = emp_df[emp_df["sku"] == sku]
            hist = round(float(row_emp["hist_avg"].sum()), 1) if not row_emp.empty else 0
            _c(ws, current_row + 1, DATA_COL_START + i,
               hist if hist > 0 else None,
               font=SMALL, fill=GREY_FILL, align=CTR, border=THIN_BRD, num_fmt=NUM_FMT1)
        _c(ws, current_row + 1, total_col, None, fill=WHITE_FILL, border=THIN_BRD)

        current_row += 2

    # ── แถว footer: รวมหีบที่กระจาย ────────────────────
    ws.row_dimensions[current_row].height = 22
    for ci in [1,2,3,4]:
        _c(ws, current_row, ci, None, fill=YELLOW_FILL, border=THIN_BRD)
    _c(ws, current_row, 5, "รวมหีบ", font=BOLD_BLK, fill=YELLOW_FILL, align=CTR, border=MED_BRD)
    foot_val = 0
    for i, sku in enumerate(skus):
        tot = int(df[df["sku"] == sku]["allocated_boxes"].sum())
        foot_val += tot * sku_price.get(sku, 0)
        off = sku_official.get(sku)
        isMatch = (tot == off) if off is not None else True
        cell_font = Font(name="Cordia New", bold=True, size=11,
                         color="2F7A4D" if isMatch else "C0392B")
        _c(ws, current_row, DATA_COL_START + i, tot,
           font=cell_font, fill=YELLOW_FILL, align=CTR, border=MED_BRD, num_fmt=NUM_FMT)
    _c(ws, current_row, total_col, foot_val,
       font=BOLD_BLK, fill=YELLOW_FILL, align=RGT, border=MED_BRD, num_fmt=BAHT_FMT)

    ws.freeze_panes = f"F{10}"
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    print(f"✅ Excel saved: {output_path} (brand={brand_filter}, emps={len(emps)}, skus={len(skus)})")
    return output_path


# ── Backward compat stubs ──────────────────────────────────
def create_mock_template(path):
    """Legacy stub — ไม่ใช้แล้ว แต่เก็บไว้ไม่ให้ import error"""
    pass

def inject_allocation_to_excel(template, result_csv, output):
    """Legacy stub — redirect ไป create_target_excel"""
    create_target_excel(result_csv, output)