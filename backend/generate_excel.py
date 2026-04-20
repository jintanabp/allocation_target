"""
generate_excel.py — สร้าง Excel กระจายเป้าหมาย
────────────────────────────────────────────────
ตาม template รูปที่ 2:
  แถว 1   : Title + supervisor info
  แถว 4   : ราคา/หีบ
  แถว 5   : รหัส SKU (bold, สีเหลือง)
  แถว 7   : เป้ารวม (supervisor_target_boxes)
  แถว 8   : เป้าย่อย (sum จากพนักงาน)
  แถว 9   : เคยขาย (hist_avg ทีม)
  แถว 10+ : พนักงาน (คู่: เป้า / เคยขาย)
  คอลัมน์ท้าย : ยอดรวม (บาท)
"""

import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ── Styles ───────────────────────────────────────────────
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
GREEN_FILL  = PatternFill("solid", fgColor="92D050")
BLUE_FILL   = PatternFill("solid", fgColor="BDD7EE")
GREY_FILL   = PatternFill("solid", fgColor="D9D9D9")
ORANGE_FILL = PatternFill("solid", fgColor="F4B942")
HDR_FILL    = PatternFill("solid", fgColor="1F497D")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
WARN_FILL   = PatternFill("solid", fgColor="FFEB9C")  # warning: ยอดเงินห่างจากเป้า

THIN = Side(style="thin", color="BFBFBF")
MED  = Side(style="medium", color="595959")
THIN_BRD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED_BRD  = Border(left=MED,  right=MED,  top=MED,  bottom=MED)

BOLD_HDR  = Font(name="Cordia New", bold=True, color="FFFFFF", size=12)
BOLD_BLK  = Font(name="Cordia New", bold=True, size=11)
NORM      = Font(name="Cordia New", size=11)
SMALL     = Font(name="Cordia New", size=10, italic=True, color="595959")
NUM_FMT   = "#,##0"
NUM_FMT1  = "#,##0.0"
BAHT_FMT  = "#,##0.00"

CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left",   vertical="center")
RGT = Alignment(horizontal="right",  vertical="center")

# Excel จำกัดชื่อชีต 31 ตัวอักษร และห้ามอักขระบางตัว
_BAD_SHEET_CHARS = '[]:*?/\\'


def _sanitize_sheet_name(name: str) -> str:
    t = str(name).strip() or "กระจายเป้า"
    for c in _BAD_SHEET_CHARS:
        t = t.replace(c, "_")
    return t[:31] if t else "Sheet"


def _ensure_unique_sheet_name(name: str, used: set[str]) -> str:
    base = _sanitize_sheet_name(name)
    if base not in used:
        used.add(base)
        return base
    for i in range(2, 1000):
        suffix = f" ({i})"
        cand = (base[: max(0, 31 - len(suffix))].rstrip() + suffix)[:31]
        if cand not in used:
            used.add(cand)
            return cand
    used.add("Sheet")
    return "Sheet"


def _brand_group_series(df: pd.DataFrame) -> pd.Series:
    """คอลัมน์ชื่อแบรนด์สำหรับจัดกลุ่มชีต (ไทยก่อน แล้วอังกฤษ แล้ว placeholder)"""
    th = df["brand_name_thai"].fillna("").astype(str).str.strip()
    en = df.get("brand_name_english", pd.Series("", index=df.index)).fillna("").astype(str).str.strip()
    out = th.copy()
    out = out.mask(out == "", en)
    out = out.mask(out == "", "(ไม่ระบุแบรนด์)")
    return out


def _emp_allocated_baht(dfx: pd.DataFrame, emp, sku_price: dict) -> float:
    """มูลค่ากระจายรวมของพนักงานใน df ชุดนี้ (บาท)"""
    sub = dfx[dfx["emp_id"] == emp]
    if sub.empty:
        return 0.0
    t = 0.0
    for sku in sub["sku"].unique():
        t += int(sub[sub["sku"] == sku]["allocated_boxes"].sum()) * float(sku_price.get(sku, 0) or 0)
    return t


def _load_sku_official(target_boxes_csv: str | None, brand_filter_for_csv: str) -> dict[str, int]:
    sku_official: dict[str, int] = {}
    tpath = target_boxes_csv or ""
    if not tpath or not os.path.exists(tpath):
        return sku_official
    try:
        df_t = pd.read_csv(tpath, dtype={"sku": str}).dropna(subset=["sku"])
        df_t["sku"] = df_t["sku"].astype(str).str.strip()
        if brand_filter_for_csv != "ALL" and "brand_name_thai" in df_t.columns:
            df_t = df_t[df_t["brand_name_thai"].astype(str) == str(brand_filter_for_csv)]
        for _, r in df_t.iterrows():
            sku_official[str(r["sku"]).strip()] = int(round(float(r.get("supervisor_target_boxes", 0) or 0)))
    except Exception as ex:
        print(f"⚠️ อ่าน {tpath} สำหรับเป้าหีบหัวหน้าไม่ได้: {ex}")
    return sku_official


def _c(ws, row, col, value=None, font=None, fill=None, align=None, border=None, num_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:    cell.font      = font
    if fill:    cell.fill      = fill
    if align:   cell.alignment = align
    if border:  cell.border    = border
    if num_fmt: cell.number_format = num_fmt
    return cell


def _write_allocation_sheet(
    ws,
    df: pd.DataFrame,
    *,
    brand_label: str,
    yellow_map: dict,
    sup_id: str,
    sku_official: dict[str, int],
    use_yellow_title_total: bool,
) -> None:
    """เขียนรูปแบบกระจายเป้าลง worksheet เดียว (df ต้องเป็น view แบรนด์เดียวหรือที่ต้องการแล้ว)"""
    # ── เรียง SKU ให้แบรนด์เดียวกันอยู่ติดกัน ───────────────────────────
    sku_brand_key = (
        df[["sku", "brand_name_thai", "brand_name_english"]]
        .astype(str)
        .drop_duplicates(subset=["sku"])
        .set_index("sku")
        .apply(
            lambda r: (str(r.get("brand_name_thai", "") or "").strip()
                       or str(r.get("brand_name_english", "") or "").strip()),
            axis=1,
        )
        .to_dict()
    )

    skus = sorted(
        df["sku"].astype(str).unique().tolist(),
        key=lambda s: (1 if (sku_brand_key.get(s, "") or "") == "" else 0, sku_brand_key.get(s, "") or "", s),
    )
    emps = df["emp_id"].unique().tolist()
    sku_price = dict(zip(df["sku"], df["price_per_box"]))
    sku_pname = {}
    for sku in skus:
        sub = df[df["sku"] == sku]
        if sub.empty:
            sku_pname[sku] = ""
            continue
        en = str(sub["product_name_english"].iloc[0] or "").strip()
        th = str(sub["product_name_thai"].iloc[0] or "").strip()
        sku_pname[sku] = en or th

    sku_allocated = df.groupby("sku")["allocated_boxes"].sum().to_dict()

    # ── Column widths ────────────────────────────────────
    ws.column_dimensions["A"].width = 5   # ลำดับ
    ws.column_dimensions["B"].width = 10  # S/M
    ws.column_dimensions["C"].width = 10  # กลุ่มย่อย
    ws.column_dimensions["D"].width = 10  # W/H
    ws.column_dimensions["E"].width = 8   # ประเภท

    DATA_COL_START = 6  # F
    for i in range(len(skus)):
        ws.column_dimensions[get_column_letter(DATA_COL_START + i)].width = 10
    total_col = DATA_COL_START + len(skus)
    ws.column_dimensions[get_column_letter(total_col)].width = 18

    # ── แถว 1: Title ────────────────────────────────────
    ws.row_dimensions[1].height = 28
    if use_yellow_title_total:
        header_total_baht = sum(yellow_map.values())
    else:
        header_total_baht = sum(
            int(df[df["sku"] == sku]["allocated_boxes"].sum()) * float(sku_price.get(sku, 0) or 0)
            for sku in skus
        )
    title_text = (
        f"ปรับปรุงเป้าหมายพนักงานขายเดือนหน้า  |  Supervisor: {sup_id}  |  "
        f"แบรนด์: {brand_label}  |  เป้ารวม: {header_total_baht:,.0f} บาท"
    )
    t = ws.cell(row=1, column=1, value=title_text)
    t.font = Font(name="Cordia New", bold=True, size=13, color="1F497D")
    t.alignment = LFT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_col)

    # ── แถว 2: ว่าง (spacer) ────────────────────────────
    ws.row_dimensions[2].height = 6

    # ── แถว 3: Header A-E + SKU group header ───────────
    ws.row_dimensions[3].height = 6

    # ── แถว 4: ราคา/หีบ ─────────────────────────────────
    ws.row_dimensions[4].height = 20
    for ci, lbl in enumerate(["", "ราคา/หีบ", "", "", ""], start=1):
        fill = GREY_FILL if lbl else WHITE_FILL
        _c(ws, 4, ci, lbl or None, font=BOLD_BLK, fill=fill, align=CTR, border=THIN_BRD)
    for i, sku in enumerate(skus):
        _c(ws, 4, DATA_COL_START + i, sku_price.get(sku, 0),
           font=BOLD_BLK, fill=YELLOW_FILL, align=CTR, border=THIN_BRD, num_fmt=NUM_FMT1)
    _c(ws, 4, total_col, None, fill=WHITE_FILL, border=THIN_BRD)

    # ── แถว 5: รหัสสินค้า SKU ───────────────────────────
    ws.row_dimensions[5].height = 22
    for ci, lbl in enumerate(["ลำดับ", "S/M", "กลุ่มย่อย", "W/H", "ประเภท"], start=1):
        _c(ws, 5, ci, lbl, font=BOLD_BLK, fill=GREY_FILL, align=CTR, border=THIN_BRD)
    for i, sku in enumerate(skus):
        _c(ws, 5, DATA_COL_START + i, sku,
           font=Font(name="Cordia New", bold=True, size=11),
           fill=YELLOW_FILL, align=CTR, border=THIN_BRD)
    _c(ws, 5, total_col, "ยอดรวม (บาท)",
       font=BOLD_HDR, fill=HDR_FILL, align=CTR, border=THIN_BRD)

    # ── แถว 6: ชื่อสินค้า (Product_NameEnglish) ───────────────────────
    ws.row_dimensions[6].height = 28
    for ci in [1, 2, 3, 4, 5]:
        _c(ws, 6, ci, None, fill=WHITE_FILL, border=THIN_BRD)
    for i, sku in enumerate(skus):
        _c(ws, 6, DATA_COL_START + i, sku_pname.get(sku, "") or None,
           font=SMALL, fill=WHITE_FILL, align=CTR, border=THIN_BRD)
    _c(ws, 6, total_col, None, fill=WHITE_FILL, border=THIN_BRD)

    # ── แถว 7: ว่าง ─────────────────────────────────────
    ws.row_dimensions[7].height = 4

    # ── แถว 8: เป้ารวม (supervisor total target boxes) ──
    ws.row_dimensions[8].height = 22
    _c(ws, 8, 1, None, fill=YELLOW_FILL, border=THIN_BRD)
    _c(ws, 8, 2, None, fill=YELLOW_FILL, border=THIN_BRD)
    _c(ws, 8, 3, None, fill=YELLOW_FILL, border=THIN_BRD)
    _c(ws, 8, 4, None, fill=YELLOW_FILL, border=THIN_BRD)
    _c(ws, 8, 5, "เป้าหีบ (หัวหน้า)", font=BOLD_BLK, fill=ORANGE_FILL, align=CTR, border=THIN_BRD)
    row7_val = 0
    for i, sku in enumerate(skus):
        boxes = int(sku_official[sku]) if sku in sku_official else int(sku_allocated.get(sku, 0))
        row7_val += boxes * sku_price.get(sku, 0)
        _c(ws, 8, DATA_COL_START + i, boxes,
           font=BOLD_BLK, fill=YELLOW_FILL, align=CTR, border=THIN_BRD, num_fmt=NUM_FMT)
    _c(ws, 8, total_col, row7_val,
       font=BOLD_BLK, fill=YELLOW_FILL, align=RGT, border=THIN_BRD, num_fmt=BAHT_FMT)

    # ── แถว 9: ผลรวมกระจาย (หลังคำนวณ) ────────────
    ws.row_dimensions[9].height = 20
    _c(ws, 9, 5, "ผลรวมกระจาย", font=BOLD_BLK, fill=YELLOW_FILL, align=CTR, border=THIN_BRD)
    for ci in [1,2,3,4]:
        _c(ws, 9, ci, None, fill=YELLOW_FILL, border=THIN_BRD)
    row8_val = 0
    for i, sku in enumerate(skus):
        boxes = int(sku_allocated.get(sku, 0))
        row8_val += boxes * sku_price.get(sku, 0)
        _c(ws, 9, DATA_COL_START + i, boxes,
           font=NORM, fill=YELLOW_FILL, align=CTR, border=THIN_BRD, num_fmt=NUM_FMT)
    _c(ws, 9, total_col, row8_val,
       font=BOLD_BLK, fill=YELLOW_FILL, align=RGT, border=THIN_BRD, num_fmt=BAHT_FMT)

    # ── แถว 10: เคยขาย (hist avg ทีม รวม) ───────────────
    ws.row_dimensions[10].height = 18
    _c(ws, 10, 5, "เคยขาย", font=SMALL, fill=GREY_FILL, align=CTR, border=THIN_BRD)
    for ci in [1,2,3,4]:
        _c(ws, 10, ci, None, fill=GREY_FILL, border=THIN_BRD)
    for i, sku in enumerate(skus):
        team_hist = df[df["sku"] == sku]["hist_avg"].sum()
        _c(ws, 10, DATA_COL_START + i, round(team_hist, 1) if team_hist > 0 else None,
           font=SMALL, fill=GREY_FILL, align=CTR, border=THIN_BRD, num_fmt=NUM_FMT1)
    _c(ws, 10, total_col, None, fill=GREY_FILL, border=THIN_BRD)

    # ── แถว 10+: รายพนักงาน ─────────────────────────────
    current_row = 11
    for idx, emp in enumerate(emps, start=1):
        ws.row_dimensions[current_row].height     = 20
        ws.row_dimensions[current_row + 1].height = 18

        emp_df = df[df["emp_id"] == emp]
        yellow_target = yellow_map.get(emp, 0)

        # คำนวณยอดรวมจริงของพนักงาน (ทุก SKU ใน view นี้)
        emp_value = sum(
            int(emp_df[emp_df["sku"] == sku]["allocated_boxes"].sum()) * sku_price.get(sku, 0)
            for sku in skus
        )
        # deviation จาก yellow_target → เลือก fill
        deviation = abs(emp_value - yellow_target)
        val_fill = GREEN_FILL if deviation <= 1000 else WARN_FILL

        # บรรทัด "เป้า"
        _c(ws, current_row, 1, idx, font=NORM, fill=BLUE_FILL, align=CTR, border=THIN_BRD)
        _c(ws, current_row, 2, emp, font=BOLD_BLK, fill=BLUE_FILL, align=CTR, border=THIN_BRD)
        _c(ws, current_row, 3, None, fill=BLUE_FILL, border=THIN_BRD)
        _c(ws, current_row, 4, None, fill=BLUE_FILL, border=THIN_BRD)
        _c(ws, current_row, 5, "เป้า", font=BOLD_BLK, fill=GREEN_FILL, align=CTR, border=THIN_BRD)
        for i, sku in enumerate(skus):
            row_emp = emp_df[emp_df["sku"] == sku]
            boxes = int(row_emp["allocated_boxes"].sum()) if not row_emp.empty else 0
            cell_fill = GREEN_FILL if boxes > 0 else WHITE_FILL
            _c(ws, current_row, DATA_COL_START + i,
               boxes if boxes > 0 else None,
               font=NORM, fill=cell_fill, align=CTR, border=THIN_BRD, num_fmt=NUM_FMT)
        _c(ws, current_row, total_col, emp_value,
           font=BOLD_BLK, fill=val_fill, align=RGT, border=THIN_BRD, num_fmt=BAHT_FMT)

        # บรรทัด "เคยขาย"
        _c(ws, current_row + 1, 1, None, fill=WHITE_FILL, border=THIN_BRD)
        _c(ws, current_row + 1, 2, emp, font=SMALL, fill=WHITE_FILL, align=CTR, border=THIN_BRD)
        _c(ws, current_row + 1, 3, None, fill=WHITE_FILL, border=THIN_BRD)
        _c(ws, current_row + 1, 4, None, fill=WHITE_FILL, border=THIN_BRD)
        _c(ws, current_row + 1, 5, "เคยขาย", font=SMALL, fill=GREY_FILL, align=CTR, border=THIN_BRD)
        for i, sku in enumerate(skus):
            row_emp = emp_df[emp_df["sku"] == sku]
            hist = round(float(row_emp["hist_avg"].sum()), 1) if not row_emp.empty else 0
            _c(ws, current_row + 1, DATA_COL_START + i,
               hist if hist > 0 else None,
               font=SMALL, fill=GREY_FILL, align=CTR, border=THIN_BRD, num_fmt=NUM_FMT1)
        _c(ws, current_row + 1, total_col, None, fill=WHITE_FILL, border=THIN_BRD)

        current_row += 2

    # ── แถว footer: รวมหีบที่กระจาย ────────────────────
    ws.row_dimensions[current_row].height = 22
    for ci in [1,2,3,4]:
        _c(ws, current_row, ci, None, fill=YELLOW_FILL, border=THIN_BRD)
    _c(ws, current_row, 5, "รวมหีบ", font=BOLD_BLK, fill=YELLOW_FILL, align=CTR, border=MED_BRD)
    foot_val = 0
    for i, sku in enumerate(skus):
        tot = int(df[df["sku"] == sku]["allocated_boxes"].sum())
        foot_val += tot * sku_price.get(sku, 0)
        off = sku_official.get(sku)
        isMatch = (tot == off) if off is not None else True
        cell_font = Font(name="Cordia New", bold=True, size=11,
                         color="2F7A4D" if isMatch else "C0392B")
        _c(ws, current_row, DATA_COL_START + i, tot,
           font=cell_font, fill=YELLOW_FILL, align=CTR, border=MED_BRD, num_fmt=NUM_FMT)
    _c(ws, current_row, total_col, foot_val,
       font=BOLD_BLK, fill=YELLOW_FILL, align=RGT, border=MED_BRD, num_fmt=BAHT_FMT)

    ws.freeze_panes = f"F{11}"


# ══════════════════════════════════════════════════════════
#  PUBLIC ENTRY POINT
# ══════════════════════════════════════════════════════════
def create_target_excel(
    result_csv:  str,
    output_path: str,
    brand_filter: str = "ALL",
    yellow_map:  dict | None = None,  # emp_id → yellow_target
    sup_id:      str = "",
    target_boxes_csv: str | None = "data/target_boxes.csv",
) -> str | None:
    """
    สร้างไฟล์ Excel จาก result_csv
    brand_filter = "ALL" → ชีตแรกรวมทุกแบรนด์ แล้วตามด้วยชีตแยกต่อแบรนด์ | "ชื่อแบรนด์" → ชีตเดียว
    yellow_map   = {emp_id: yellow_target_baht} สำหรับแสดงใน header และ validate deviation
    target_boxes_csv — ถ้ามีไฟล์ จะอ่าน supervisor_target_boxes มาแสดงเป็นแถว "เป้าหีบ (หัวหน้า)"
    แถวถัดไปเป็น "ผลรวมกระจาย" เพื่อให้เทียบกับผลคำนวณได้
    """
    if not os.path.exists(result_csv):
        print(f"❌ ไม่พบ {result_csv}")
        return None

    df = pd.read_csv(result_csv, dtype={"sku": str, "emp_id": str})
    df["allocated_boxes"] = pd.to_numeric(df["allocated_boxes"], errors="coerce").fillna(0).astype(int)
    df["hist_avg"]        = pd.to_numeric(df["hist_avg"],        errors="coerce").fillna(0.0)
    df["hist_ly_same_month"] = pd.to_numeric(
        df.get("hist_ly_same_month", 0), errors="coerce"
    ).fillna(0.0)
    df["hist_prev_month"] = pd.to_numeric(
        df.get("hist_prev_month", 0), errors="coerce"
    ).fillna(0.0)
    df["price_per_box"]   = pd.to_numeric(df["price_per_box"],   errors="coerce").fillna(0.0)
    df["brand_name_thai"] = df.get("brand_name_thai", pd.Series("", index=df.index)).fillna("").astype(str)
    df["product_name_thai"] = df.get("product_name_thai", pd.Series("", index=df.index)).fillna("").astype(str)
    df["product_name_english"] = df.get("product_name_english", pd.Series("", index=df.index)).fillna("").astype(str)
    df["brand_name_english"] = df.get("brand_name_english", pd.Series("", index=df.index)).fillna("").astype(str)

    if brand_filter != "ALL":
        df = df[df["brand_name_thai"] == brand_filter].copy()
        if df.empty:
            print(f"⚠️ ไม่พบข้อมูลสำหรับแบรนด์ '{brand_filter}'")
            return None

    if df.empty:
        print("⚠️ ไม่มีข้อมูล allocation")
        return None

    yellow_map = yellow_map or {}
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)

    if brand_filter == "ALL":
        df = df.copy()
        df["_brand_grp"] = _brand_group_series(df)
        brands = sorted(df["_brand_grp"].unique().tolist())
        sku_price_full = dict(zip(df["sku"], df["price_per_box"]))
        emp_totals = {
            e: _emp_allocated_baht(df, e, sku_price_full)
            for e in df["emp_id"].unique()
        }
        wb = openpyxl.Workbook()
        used_sheet_names: set[str] = set()

        df_all = df.drop(columns=["_brand_grp"], errors="ignore")
        sku_off_all = _load_sku_official(target_boxes_csv, "ALL")
        ws_all = wb.active
        ws_all.title = _ensure_unique_sheet_name("รวมทุกแบรนด์", used_sheet_names)
        _write_allocation_sheet(
            ws_all,
            df_all,
            brand_label="ทุกแบรนด์",
            yellow_map=yellow_map,
            sup_id=sup_id,
            sku_official=sku_off_all,
            use_yellow_title_total=True,
        )

        for b in brands:
            df_b = df[df["_brand_grp"] == b].drop(columns=["_brand_grp"], errors="ignore")
            if df_b.empty:
                continue
            sku_off = _load_sku_official(target_boxes_csv, b)
            sheet_name = _ensure_unique_sheet_name(b, used_sheet_names)
            ws = wb.create_sheet(sheet_name)
            yellow_map_brand: dict = {}
            for e in df_b["emp_id"].unique():
                tot = emp_totals.get(e, 0.0)
                sl = _emp_allocated_baht(df_b, e, sku_price_full)
                base_y = float(yellow_map.get(e, 0) or 0)
                yellow_map_brand[e] = base_y * (sl / tot) if tot > 0 else 0.0
            _write_allocation_sheet(
                ws,
                df_b,
                brand_label=b,
                yellow_map=yellow_map_brand,
                sup_id=sup_id,
                sku_official=sku_off,
                use_yellow_title_total=False,
            )
        wb.save(output_path)
        print(f"✅ Excel saved: {output_path} (brand=ALL, sheets={len(used_sheet_names)})")
        return output_path

    sku_off = _load_sku_official(target_boxes_csv, brand_filter)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "กระจายเป้า"
    _write_allocation_sheet(
        ws,
        df,
        brand_label=brand_filter,
        yellow_map=yellow_map,
        sup_id=sup_id,
        sku_official=sku_off,
        # export แบรนด์เดียว: ให้ "เป้ารวม" เป็นยอดของแบรนด์นี้ (ไม่ใช่ยอดเหลืองรวมทุกแบรนด์)
        use_yellow_title_total=False,
    )
    wb.save(output_path)
    n_emp = df["emp_id"].nunique()
    n_sku = df["sku"].nunique()
    print(f"✅ Excel saved: {output_path} (brand={brand_filter}, emps={n_emp}, skus={n_sku})")
    return output_path


# ── Backward compat stubs ──────────────────────────────────
def create_mock_template(path):
    """Legacy stub — ไม่ใช้แล้ว แต่เก็บไว้ไม่ให้ import error"""
    pass

def inject_allocation_to_excel(template, result_csv, output):
    """Legacy stub — redirect ไป create_target_excel"""
    create_target_excel(result_csv, output)