import io
import logging
import os
import uuid
from datetime import datetime, timezone
from zoneinfo import ZoneInfo

import msal
import pandas as pd
import requests
from fastapi import HTTPException
from openpyxl import Workbook

from ..core.paths import safe_id, tga_grain_cache_path
from ..fabric_dax_connector import FabricDAXConnector
from ..schemas import LakehouseUploadRequest

logger = logging.getLogger("target_allocation")

LAKEHOUSE_TEXT_DATE_COLUMNS = frozenset({"EFFECTIVEDATE", "UPDATEDATE"})

LAKEHOUSE_CSV_COLUMNS = [
    "PRODUCTCODE",
    "SALESTYPE",
    "DIVISIONCODE",
    "SALESMANCODE",
    "AREACODE",
    "PROVINCECODE",
    "WAREHOUSECODE",
    "QUANTITYCASE",
    "EFFECTIVEDATE",
    "UPDATEDATE",
    "USERCODE",
]


def _get_storage_token() -> str:
    tenant_id = (os.environ.get("FABRIC_TENANT_ID") or "").strip()
    client_id = (os.environ.get("FABRIC_CLIENT_ID") or "").strip()
    client_secret = (os.environ.get("FABRIC_CLIENT_SECRET") or "").strip()
    if not (tenant_id and client_id and client_secret):
        raise HTTPException(
            500,
            detail=(
                "ยังไม่ได้ตั้งค่า Service Principal สำหรับอัปโหลดเข้า OneLake "
                "(ต้องมี FABRIC_TENANT_ID / FABRIC_CLIENT_ID / FABRIC_CLIENT_SECRET)"
            ),
        )

    authority = f"https://login.microsoftonline.com/{tenant_id}"
    cca = msal.ConfidentialClientApplication(
        client_id,
        client_credential=client_secret,
        authority=authority,
    )
    scopes = ["https://storage.azure.com/.default"]
    result = cca.acquire_token_for_client(scopes=scopes)
    token = result.get("access_token")
    if token:
        return token
    err = result.get("error_description") or result.get("error") or str(result)
    raise HTTPException(500, detail=f"ขอ token สำหรับอัปโหลด OneLake ไม่สำเร็จ: {err}")


def _onelake_base_path() -> tuple[str, str]:
    ws = (os.environ.get("ONELAKE_WORKSPACE_ID") or "").strip()
    lh = (os.environ.get("ONELAKE_LAKEHOUSE_ID") or "").strip()
    if not ws or not lh:
        raise HTTPException(
            500,
            detail=(
                "ยังไม่ได้ตั้งค่าเป้าหมาย Lakehouse (ต้องมี ONELAKE_WORKSPACE_ID / ONELAKE_LAKEHOUSE_ID)"
            ),
        )
    return ws, lh


def _onelake_file_url(file_path: str) -> tuple[str, str]:
    ws, lh = _onelake_base_path()
    base = "https://onelake.dfs.fabric.microsoft.com"
    fp = file_path.lstrip("/").replace("\\", "/")
    if fp.lower().startswith("files/"):
        fp = fp[6:]
    return f"{base}/{ws}/{lh}/Files/{fp}", fp


def _onelake_delete_if_exists(url: str, headers: dict) -> None:
    r = requests.delete(url, headers=headers, timeout=60)
    if r.status_code in (200, 202, 404):
        return
    logger.warning(
        "OneLake delete before upload: HTTP %s — %s", r.status_code, (r.text or "")[:200]
    )


def _bangkok_date_yyyymmdd() -> str:
    return datetime.now(ZoneInfo("Asia/Bangkok")).strftime("%Y%m%d")


def _format_datetime_bangkok_be(dt: datetime) -> str:
    """รูปแบบ d/M/yyyy HH:mm:ss ปี พ.ศ. แบบ 24 ชม. (ไม่มี AM/PM)"""
    return (
        f"{dt.day}/{dt.month}/{dt.year + 543} "
        f"{dt.hour:02d}:{dt.minute:02d}:{dt.second:02d}"
    )


def _format_updatedate_bangkok_be() -> str:
    return _format_datetime_bangkok_be(datetime.now(ZoneInfo("Asia/Bangkok")))


def _format_effectivedate_bangkok_be(target_year: int, target_month: int) -> str:
    """วันแรกของเดือนเป้า เวลา 00:00:00 (ปฏิทิน พ.ศ.)"""
    dt = datetime(
        int(target_year),
        int(target_month),
        1,
        0,
        0,
        0,
        tzinfo=ZoneInfo("Asia/Bangkok"),
    )
    return _format_datetime_bangkok_be(dt)


def _cell_str(val) -> str:
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    return str(val).strip()


def _areacode_str(val) -> str:
    """
    ค่า AREACODE ใน semantic model — รวม **0** — ต้องส่งออกตามนั้น เพื่อ import กลับ TGA
    เดิมเคยตัด 0 ออกเป็น "" (ถือว่าว่าง) ทำให้ดูเหมือนโมเดลไม่มีรหัสพื้นที่แม้ฐานข้อมูลเป็น 0
    """
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        if float(val) == int(val):
            return str(int(val))
        return str(val).strip()
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none"):
        return ""
    low = s.lower()
    if low in ("0", "0.0", "-0", "-0.0"):
        return "0"
    return s


def _resolve_user_code(req: LakehouseUploadRequest) -> str:
    """
    รหัสผู้บันทึก: ส่งจาก frontend (manager หรือ supervisor ที่ล็อกอิน)
    สำรองเป็น sup_id ของทีมที่กำลังเกลี่ย
    """
    if req.upload_user_code and str(req.upload_user_code).strip():
        return str(req.upload_user_code).strip().upper()
    return str(req.sup_id or "").strip().upper()


def _coalesce_col(df: pd.DataFrame, col: str, fallback: pd.Series | None = None) -> pd.Series:
    base = df[col].map(_cell_str) if col in df.columns else pd.Series([""] * len(df), index=df.index)
    if fallback is not None:
        return base.where(base.ne(""), fallback.map(_cell_str))
    return base


def _integer_split_by_weights(weights: list[float], total: int) -> list[int]:
    """หีบแบ่งจำนวนเต็มผลรวม total ตาม weights (เกลี่ยเศษตาม leftover มากที่สุด)"""
    total = max(0, int(round(total)))
    n = len(weights)
    if n == 0:
        return []
    if total == 0:
        return [0] * n
    w = [max(0.0, float(x)) for x in weights]
    s = sum(w)
    if s <= 0:
        w = [1.0] * n
        s = float(n)
    raw = [total * (wi / s) for wi in w]
    base = [int(x) for x in raw]
    rem = total - sum(base)
    frac = sorted(range(n), key=lambda i: -(raw[i] - base[i]))
    for j in range(rem):
        base[frac[j % n]] += 1
    return base


def _normalize_grain_dtype(df_grain: pd.DataFrame) -> pd.DataFrame:
    g = df_grain.copy()
    for c in ("emp_id", "sku"):
        if c in g.columns:
            g[c] = g[c].astype(str).str.strip()
    for c in ("salestype", "divisioncode", "areacode", "provincecode", "warehouse_code"):
        if c not in g.columns:
            g[c] = ""
        else:
            g[c] = g[c].map(_cell_str)
    if "qty" not in g.columns:
        g["qty"] = 0.0
    g["qty"] = pd.to_numeric(g["qty"], errors="coerce").fillna(0.0)
    return g


def _expand_allocations_with_tga_grain(
    df_alloc: pd.DataFrame,
    sup_id: str,
    target_month: int,
    target_year: int,
) -> tuple[pd.DataFrame, bool]:
    """
    จากแถว (emp×sku × allocated_boxes) แตกเป็นหลายแถวตาม grain cache จาก tga_target_salesman_next
    ให้ SALESTYPE / DIVISIONCODE / AREACODE / PROVINCECODE / WAREHOUSECODE ตรงกับบรรทัดเป้า
    และรักษายอด QUANTITYCASE รวมต่อ emp×sku
    """
    p = tga_grain_cache_path(sup_id, target_month, target_year)
    if not os.path.exists(p):
        return df_alloc, False
    try:
        dg = pd.read_csv(p, dtype=str, keep_default_na=False)
    except Exception:
        logger.warning("cannot read TGA grain cache: %s", p)
        return df_alloc, False
    dg = _normalize_grain_dtype(dg)
    if dg.empty:
        return df_alloc, False

    out: list[dict] = []

    # แปลง DataFrame iterate — เก็บ warehouse จากคำขอเป็น hint
    for _, arow in df_alloc.iterrows():
        e = str(arow["emp_id"]).strip()
        sku = str(arow["sku"]).strip()
        boxes_val = pd.to_numeric(arow.get("allocated_boxes", 0), errors="coerce")
        boxes = 0 if pd.isna(boxes_val) else int(round(float(boxes_val)))
        wh_req = ""
        raw_wh = arow.get("warehouse_code")
        if raw_wh is not None and str(raw_wh).strip():
            wh_req = str(raw_wh).strip()

        sub = dg[(dg["emp_id"] == e) & (dg["sku"] == sku)].copy()

        # ไม่พบใน cache → เก็บบรรทัดเดิมให้ชั้นถัดไปเติม dim จาก Fabric
        if sub.empty:
            out.append(
                {
                    "emp_id": e,
                    "sku": sku,
                    "allocated_boxes": boxes,
                    "salestype": "",
                    "divisioncode": "",
                    "areacode": "",
                    "provincecode": "",
                    "warehouse_code": wh_req,
                }
            )
            continue

        sub_pos = sub[sub["qty"] > 0]
        dims_only = sub[sub["qty"] <= 0]

        if not sub_pos.empty:
            wvals = sub_pos["qty"].astype(float).tolist()
            split = _integer_split_by_weights(wvals, boxes)
            for (_, r), b in zip(sub_pos.iterrows(), split):
                wh = (
                    _cell_str(r.get("warehouse_code", ""))
                    or wh_req
                    or ""
                )
                out.append(
                    {
                        "emp_id": e,
                        "sku": sku,
                        "allocated_boxes": int(b),
                        "salestype": _cell_str(r.get("salestype", "")),
                        "divisioncode": _cell_str(r.get("divisioncode", "")),
                        "areacode": _areacode_str(r.get("areacode", "")),
                        "provincecode": _cell_str(r.get("provincecode", "")),
                        "warehouse_code": wh,
                    }
                )

            # แถว TGA เดิมที่ qty = 0: เขียน QUANTITYCASE = 0 เพื่อให้ครบ dim ตอนนำเข้ากลับ
            if not dims_only.empty:
                for _, r in dims_only.iterrows():
                    wh = (
                        _cell_str(r.get("warehouse_code", ""))
                        or wh_req
                        or ""
                    )
                    out.append(
                        {
                            "emp_id": e,
                            "sku": sku,
                            "allocated_boxes": 0,
                            "salestype": _cell_str(r.get("salestype", "")),
                            "divisioncode": _cell_str(r.get("divisioncode", "")),
                            "areacode": _areacode_str(r.get("areacode", "")),
                            "provincecode": _cell_str(r.get("provincecode", "")),
                            "warehouse_code": wh,
                        }
                    )
        else:
            # เฉพาะแถวเป้ารวมเป็น 0 ใน TGA → เกลี่ยหีบเท่า ๆ กันบนทุกความเป็นไปได้ของ dim
            wvals = [1.0] * len(sub)
            split = _integer_split_by_weights(wvals, boxes)
            for (_, r), b in zip(sub.iterrows(), split):
                wh = (
                    _cell_str(r.get("warehouse_code", ""))
                    or wh_req
                    or ""
                )
                out.append(
                    {
                        "emp_id": e,
                        "sku": sku,
                        "allocated_boxes": int(b),
                        "salestype": _cell_str(r.get("salestype", "")),
                        "divisioncode": _cell_str(r.get("divisioncode", "")),
                        "areacode": _areacode_str(r.get("areacode", "")),
                        "provincecode": _cell_str(r.get("provincecode", "")),
                        "warehouse_code": wh,
                    }
                )

    return pd.DataFrame(out), True


def _team_emp_and_sku_lists(
    df: pd.DataFrame,
    sup_id: str,
    target_month: int,
    target_year: int,
) -> tuple[list[str], list[str]]:
    """รายชื่อพนักงาน/SKU ครบทีม — ไม่พึ่งแถวที่ OR engine ส่งมา (มีแต่หีบ > 0)"""
    emps: set[str] = set()
    skus: set[str] = set()
    if not df.empty:
        emps.update(str(e).strip() for e in df["emp_id"].unique() if str(e).strip())
        skus.update(str(s).strip() for s in df["sku"].unique() if str(s).strip())

    for path, emp_col, sku_col in (
        ("data/target_sun.csv", "emp_id", None),
        ("data/target_boxes.csv", None, "sku"),
    ):
        if not os.path.exists(path):
            continue
        try:
            part = pd.read_csv(path, dtype=str)
        except Exception:
            continue
        if emp_col and emp_col in part.columns:
            emps.update(part[emp_col].astype(str).str.strip())
        if sku_col and sku_col in part.columns:
            skus.update(part[sku_col].astype(str).str.strip())

    p = tga_grain_cache_path(sup_id, target_month, target_year)
    if os.path.exists(p):
        try:
            dg = pd.read_csv(p, dtype=str, keep_default_na=False)
            if "emp_id" in dg.columns:
                emps.update(dg["emp_id"].astype(str).str.strip())
            if "sku" in dg.columns:
                skus.update(dg["sku"].astype(str).str.strip())
        except Exception:
            pass

    emps.discard("")
    skus.discard("")
    return sorted(emps), sorted(skus)


def _expand_to_team_full_matrix(
    df: pd.DataFrame,
    sup_id: str,
    target_month: int,
    target_year: int,
) -> pd.DataFrame:
    """
    เติมทุกคู่ emp×sku ของทีมด้วย allocated_boxes=0 ถ้ายังไม่มีใน payload
    (OR engine / draft เก่ามักส่งเฉพาะหีบ > 0 ทำให้ Excel ไม่มีแถว 0)
    """
    emps, skus = _team_emp_and_sku_lists(df, sup_id, target_month, target_year)
    if not emps or not skus:
        return df

    wh_by_emp: dict[str, str] = {}
    lookup: dict[tuple[str, str], int] = {}
    for _, r in df.iterrows():
        e = str(r["emp_id"]).strip()
        sku = str(r["sku"]).strip()
        if not e or not sku:
            continue
        wh = _cell_str(r.get("warehouse_code", ""))
        if wh:
            wh_by_emp[e] = wh
        boxes = pd.to_numeric(r.get("allocated_boxes", 0), errors="coerce")
        boxes = 0 if pd.isna(boxes) else int(round(float(boxes)))
        key = (e, sku)
        lookup[key] = lookup.get(key, 0) + boxes

    rows: list[dict] = []
    for e in emps:
        for sku in skus:
            rows.append(
                {
                    "emp_id": e,
                    "sku": sku,
                    "allocated_boxes": int(lookup.get((e, sku), 0)),
                    "warehouse_code": wh_by_emp.get(e, ""),
                }
            )
    out = pd.DataFrame(rows)
    logger.info(
        "lakehouse full matrix: %d emp × %d sku = %d rows (zeros=%d)",
        len(emps),
        len(skus),
        len(out),
        int((out["allocated_boxes"] == 0).sum()),
    )
    return out


def _zero_sum_emp_sku_pairs(df: pd.DataFrame) -> set[tuple[str, str]]:
    if df.empty:
        return set()
    gsum = df.groupby(["emp_id", "sku"], as_index=False)["allocated_boxes"].sum()
    return {
        (str(r.emp_id).strip(), str(r.sku).strip())
        for _, r in gsum.iterrows()
        if int(r.allocated_boxes) == 0
    }


def _load_tga_grain_frame(
    sup_id: str,
    target_month: int,
    target_year: int,
    emp_list: list[str],
) -> pd.DataFrame:
    """อ่าน grain จาก cache ขั้น Step 1; ถ้าไม่พอค่อยดึง Fabric สด"""
    p = tga_grain_cache_path(sup_id, target_month, target_year)
    if os.path.exists(p):
        try:
            dg = pd.read_csv(p, dtype=str, keep_default_na=False)
            dg = _normalize_grain_dtype(dg)
            if emp_list:
                emps = {str(e).strip() for e in emp_list}
                dg = dg[dg["emp_id"].isin(emps)]
            if not dg.empty:
                return dg
        except Exception as e:
            logger.warning("read tga grain cache for zero align: %s", e)
    try:
        fabric = FabricDAXConnector()
        return fabric.get_tga_target_salesman_granular(emp_list, target_month, target_year)
    except Exception as e:
        logger.warning("fabric granular for zero align: %s", e)
        return pd.DataFrame()


def _align_zero_allocations_to_tga_grain(
    df: pd.DataFrame,
    sup_id: str,
    target_month: int,
    target_year: int,
) -> pd.DataFrame:
    """
    คู่ emp×sku ที่หีบรวม = 0 ต้องส่งแถวที่ key ตรง Oracle (SALESTYPE+DIVISION+AREA+PROVINCE…)
    มิฉะนั้น import จะ insert แถวใหม่ qty=0 แต่แถวเดิม qty=1 ยังค้าง
    """
    zero_pairs = _zero_sum_emp_sku_pairs(df)
    if not zero_pairs:
        return df

    emp_list = sorted({e for e, _ in zero_pairs})
    dg = _load_tga_grain_frame(sup_id, target_month, target_year, emp_list)
    if dg.empty:
        logger.warning(
            "zero allocations: ไม่มี TGA grain สำหรับ %d คู่ emp×sku — "
            "อาจอัปเดต Oracle ไม่ครบ (โหลด Step 1 ใหม่ก่อนส่ง)",
            len(zero_pairs),
        )
        return df

    dg = _normalize_grain_dtype(dg)
    mask_keep = ~df.apply(
        lambda r: (str(r["emp_id"]).strip(), str(r["sku"]).strip()) in zero_pairs,
        axis=1,
    )
    kept = df[mask_keep].copy()
    zero_rows: list[dict] = []
    missing_grain: list[tuple[str, str]] = []

    for e, sku in sorted(zero_pairs):
        sub = dg[(dg["emp_id"] == e) & (dg["sku"] == sku)]
        hint = df[(df["emp_id"] == e) & (df["sku"] == sku)]
        wh_hint = _cell_str(hint.iloc[0].get("warehouse_code", "")) if not hint.empty else ""
        if sub.empty:
            missing_grain.append((e, sku))
            if not hint.empty:
                zero_rows.extend(hint.to_dict(orient="records"))
            continue
        for _, r in sub.iterrows():
            zero_rows.append(
                {
                    "emp_id": e,
                    "sku": sku,
                    "allocated_boxes": 0,
                    "salestype": _cell_str(r.get("salestype", "")),
                    "divisioncode": _cell_str(r.get("divisioncode", "")),
                    "areacode": _areacode_str(r.get("areacode", "")),
                    "provincecode": _cell_str(r.get("provincecode", "")),
                    "warehouse_code": _cell_str(r.get("warehouse_code", "")) or wh_hint,
                }
            )

    if missing_grain:
        logger.warning(
            "zero allocations without TGA grain rows: %s",
            missing_grain[:10],
        )

    if not zero_rows:
        return df
    return pd.concat([kept, pd.DataFrame(zero_rows)], ignore_index=True)


def _ensure_zero_pairs_have_rows(df: pd.DataFrame, zero_pairs: set[tuple[str, str]]) -> pd.DataFrame:
    """กันคู่ emp×sku ที่หีบ=0 หลุดจาก export เมื่อไม่มี grain ใน TGA"""
    if not zero_pairs:
        return df
    present = {
        (str(r["emp_id"]).strip(), str(r["sku"]).strip())
        for _, r in df.iterrows()
    }
    extra: list[dict] = []
    for e, sku in sorted(zero_pairs):
        if (e, sku) in present:
            continue
        extra.append(
            {
                "emp_id": e,
                "sku": sku,
                "allocated_boxes": 0,
                "salestype": "",
                "divisioncode": "",
                "areacode": "",
                "provincecode": "",
                "warehouse_code": "",
            }
        )
    if not extra:
        return df
    logger.warning(
        "added %d zero emp×sku rows missing after grain align",
        len(extra),
    )
    return pd.concat([df, pd.DataFrame(extra)], ignore_index=True)


def _enrich_emp_dimensions(
    df: pd.DataFrame,
    rows_raw: list[dict],
    skip_emp_sku_dim_merge: bool = False,
) -> pd.DataFrame:
    emp_list = sorted({str(e).strip() for e in df["emp_id"].unique() if str(e).strip()})
    sku_list = sorted({str(s).strip() for s in df["sku"].unique() if str(s).strip()})
    wh_hint = {}
    for r in rows_raw:
        emp = str(r.get("emp_id") or "").strip()
        wh = _cell_str(r.get("warehouse_code"))
        if emp and wh:
            wh_hint[emp] = wh

    df_es = pd.DataFrame()
    df_emp = pd.DataFrame()
    df_wh = pd.DataFrame()
    try:
        fabric = FabricDAXConnector()
        if not skip_emp_sku_dim_merge:
            try:
                df_es = fabric.get_tga_lakehouse_dims_by_emp_sku(emp_list, sku_list)
            except Exception as e:
                logger.warning("get_tga_lakehouse_dims_by_emp_sku: %s", e)
        try:
            df_emp = fabric.get_tga_lakehouse_dims_by_emp(emp_list)
        except Exception as e:
            logger.warning("get_tga_lakehouse_dims_by_emp: %s", e)
        try:
            df_wh = fabric.get_warehouse_by_emp(emp_list)
        except Exception as e:
            logger.warning("get_warehouse_by_emp (lakehouse): %s", e)
    except Exception as e:
        logger.warning("Fabric connector (lakehouse enrich): %s", e)

    for c in ("salestype", "divisioncode", "areacode", "provincecode", "warehouse_code"):
        if c not in df.columns:
            df[c] = ""

    if not skip_emp_sku_dim_merge and not df_es.empty:
        df = df.merge(df_es, on=["emp_id", "sku"], how="left", suffixes=("", "_tga"))

    emp_fb = {}
    if not df_emp.empty:
        emp_fb = df_emp.set_index("emp_id").to_dict(orient="index")

    def _emp_fb_series(col: str) -> pd.Series:
        if not emp_fb:
            return pd.Series([""] * len(df), index=df.index)
        return df["emp_id"].map(lambda e: _cell_str((emp_fb.get(str(e).strip()) or {}).get(col)))

    df["salestype"] = _coalesce_col(df, "salestype", _emp_fb_series("salestype"))
    df["divisioncode"] = _coalesce_col(df, "divisioncode", _emp_fb_series("divisioncode"))
    df["areacode"] = _coalesce_col(df, "areacode", _emp_fb_series("areacode"))
    df["provincecode"] = _coalesce_col(df, "provincecode", _emp_fb_series("provincecode"))

    if not df_wh.empty:
        df = df.merge(
            df_wh.rename(columns={"warehouse_code": "warehouse_hist"}),
            on="emp_id",
            how="left",
        )
        if "warehouse_code" not in df.columns:
            df["warehouse_code"] = ""
        df["warehouse_code"] = df.apply(
            lambda row: _cell_str(row.get("warehouse_code"))
            or _cell_str(row.get("warehouse_hist"))
            or wh_hint.get(str(row["emp_id"]).strip(), ""),
            axis=1,
        )
        if "warehouse_hist" in df.columns:
            df = df.drop(columns=["warehouse_hist"])
    else:
        df["warehouse_code"] = df.apply(
            lambda row: _cell_str(row.get("warehouse_code"))
            or wh_hint.get(str(row["emp_id"]).strip(), ""),
            axis=1,
        )

    wh_tga = _coalesce_col(df, "warehouse_code", _emp_fb_series("warehouse_code"))
    df["warehouse_code"] = wh_tga
    df["areacode"] = df["areacode"].map(_areacode_str)
    df["warehouse_code"] = df["warehouse_code"].map(_cell_str)
    return df


def _build_tga_upload_dataframe(req: LakehouseUploadRequest) -> pd.DataFrame:
    rows_raw = [a.model_dump() for a in req.allocations]
    df = pd.DataFrame(rows_raw)
    df["allocated_boxes"] = pd.to_numeric(df["allocated_boxes"], errors="coerce").fillna(0).astype(int)
    if df.empty:
        raise HTTPException(400, detail="ไม่มีข้อมูล allocations สำหรับส่งออก")

    df["emp_id"] = df["emp_id"].astype(str).str.strip()
    df["sku"] = df["sku"].astype(str).str.strip()
    df = df[(df["emp_id"] != "") & (df["sku"] != "")].copy()
    if df.empty:
        raise HTTPException(400, detail="ไม่มีแถว emp×sku ที่สมบูรณ์สำหรับส่งออก")

    df = _expand_to_team_full_matrix(
        df,
        req.sup_id,
        int(req.target_month),
        int(req.target_year),
    )
    zero_pairs_full = _zero_sum_emp_sku_pairs(df)

    df_expand, grain_ok = _expand_allocations_with_tga_grain(
        df,
        req.sup_id,
        int(req.target_month),
        int(req.target_year),
    )
    df = df_expand if grain_ok else df
    df = _align_zero_allocations_to_tga_grain(
        df,
        req.sup_id,
        int(req.target_month),
        int(req.target_year),
    )
    df = _ensure_zero_pairs_have_rows(df, zero_pairs_full)
    df = _enrich_emp_dimensions(
        df, rows_raw, skip_emp_sku_dim_merge=bool(grain_ok)
    )

    user_code = _resolve_user_code(req)
    updatedate = _format_updatedate_bangkok_be()
    effectivedate = _format_effectivedate_bangkok_be(req.target_year, req.target_month)

    out = pd.DataFrame(
        {
            "PRODUCTCODE": df["sku"],
            "SALESTYPE": df["salestype"].map(_cell_str),
            "DIVISIONCODE": df["divisioncode"].map(_cell_str),
            "SALESMANCODE": df["emp_id"],
            "AREACODE": df["areacode"].map(_areacode_str),
            "PROVINCECODE": df["provincecode"].map(_cell_str),
            "WAREHOUSECODE": df["warehouse_code"].map(_cell_str),
            "QUANTITYCASE": df["allocated_boxes"].astype(int),
            "EFFECTIVEDATE": effectivedate,
            "UPDATEDATE": updatedate,
            "USERCODE": user_code,
        }
    )
    return out[LAKEHOUSE_CSV_COLUMNS]


def _export_basename(req: LakehouseUploadRequest) -> str:
    day_tag = _bangkok_date_yyyymmdd()
    return f"alloc_{safe_id(req.sup_id)}_{req.target_year}_{req.target_month:02d}_{day_tag}"


def prepare_lakehouse_csv(req: LakehouseUploadRequest) -> tuple[bytes, str, pd.DataFrame]:
    """CSV สำหรับ ingest / OneLake (ค่าวันที่เป็นข้อความ d/M/yyyy HH:mm:ss)"""
    df = _build_tga_upload_dataframe(req)
    buf = io.StringIO()
    df.to_csv(buf, index=False)
    content = ("\ufeff" + buf.getvalue()).encode("utf-8")
    return content, f"{_export_basename(req)}.csv", df


def prepare_lakehouse_xlsx(req: LakehouseUploadRequest) -> tuple[bytes, str, pd.DataFrame]:
    """
    Excel สำหรับเปิดดู/แก้ — คอลัมน์วันที่เป็นข้อความ (@) เลี่ยง Excel แปลงเป็น 12:00 AM
    """
    df = _build_tga_upload_dataframe(req)
    wb = Workbook()
    ws = wb.active
    ws.title = "TGA"
    ws.append(LAKEHOUSE_CSV_COLUMNS)
    col_idx = {name: i + 1 for i, name in enumerate(LAKEHOUSE_CSV_COLUMNS)}
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    for r in range(2, ws.max_row + 1):
        for name in LAKEHOUSE_TEXT_DATE_COLUMNS:
            cell = ws.cell(row=r, column=col_idx[name])
            cell.number_format = "@"
            if cell.value is not None:
                cell.value = str(cell.value)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), f"{_export_basename(req)}.xlsx", df


def _upload_bytes_to_onelake(file_path: str, content: bytes, token: str) -> None:
    url, _fp = _onelake_file_url(file_path)

    headers = {
        "Authorization": f"Bearer {token}",
        "x-ms-version": "2021-08-06",
    }

    _onelake_delete_if_exists(url, headers)

    r0 = requests.put(url + "?resource=file", headers=headers, timeout=60)
    if r0.status_code not in (201, 200, 202):
        raise HTTPException(
            502,
            detail=f"สร้างไฟล์บน OneLake ไม่สำเร็จ (HTTP {r0.status_code}): {r0.text[:300]}",
        )

    r1 = requests.patch(
        url + "?action=append&position=0",
        headers={**headers, "Content-Type": "application/octet-stream"},
        data=content,
        timeout=120,
    )
    if r1.status_code not in (202, 200):
        raise HTTPException(
            502,
            detail=f"อัปโหลดเนื้อหาไป OneLake ไม่สำเร็จ (HTTP {r1.status_code}): {r1.text[:300]}",
        )

    r2 = requests.patch(
        url + f"?action=flush&position={len(content)}",
        headers=headers,
        timeout=60,
    )
    if r2.status_code not in (200, 201):
        raise HTTPException(
            502,
            detail=f"ยืนยันไฟล์ (flush) บน OneLake ไม่สำเร็จ (HTTP {r2.status_code}): {r2.text[:300]}",
        )


def export_allocations_excel(req: LakehouseUploadRequest) -> dict:
    """สร้าง Excel รูปแบบ tga_target_salesman_next — รวม QUANTITYCASE=0 สำหรับทับข้อมูลเดิม"""
    if not req.allocations:
        raise HTTPException(400, detail="ไม่มีข้อมูล allocations สำหรับส่งออก")

    content, fname, df = prepare_lakehouse_xlsx(req)
    zero_rows = int((df["QUANTITYCASE"] == 0).sum())
    return {
        "content": content,
        "filename": fname,
        "rows": int(len(df)),
        "zero_rows": zero_rows,
        "columns": LAKEHOUSE_CSV_COLUMNS,
    }


def upload_allocations_to_lakehouse(req: LakehouseUploadRequest) -> dict:
    """อัปโหลด CSV ไป OneLake (ใช้เมื่อเปิด ingest อัตโนมัติในอนาคต)"""
    content, fname, df = prepare_lakehouse_csv(req)
    batch_id = str(uuid.uuid4())
    uploaded_at = datetime.now(timezone.utc).isoformat()

    prefix = (os.environ.get("ONELAKE_UPLOAD_DIR") or "Files/target_allocation_uploads").strip()
    prefix = prefix.strip("/").replace("\\", "/")
    if prefix.lower().startswith("files/"):
        prefix = prefix[6:]

    remote_path = f"{prefix}/{fname}"

    token = _get_storage_token()
    _upload_bytes_to_onelake(remote_path, content, token)

    logger.info(
        "uploaded TGA-format allocations to OneLake: %s (%d rows) batch=%s",
        remote_path,
        len(df),
        batch_id,
    )
    return {
        "status": "ok",
        "rows": int(len(df)),
        "remote_path": remote_path,
        "upload_batch_id": batch_id,
        "uploaded_at_utc": uploaded_at,
        "columns": LAKEHOUSE_CSV_COLUMNS,
    }
